from flask import Flask, render_template, request, redirect, session, send_file
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from uuid import uuid4
import os
import json
import boto3
import calendar
from botocore.exceptions import ClientError

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_secret_key")

app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    "sqlite:///dkss.db"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
class Company(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(
        db.String(50),
        unique=True,
        nullable=False
    )

    company_name = db.Column(
        db.String(100),
        nullable=False
    )

    vehicle_limit = db.Column(
        db.Integer,
        default=0
    )

    active = db.Column(
        db.Boolean,
        default=True
    )


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(
        db.String(50),
        nullable=False
    )

    username = db.Column(
        db.String(50),
        nullable=False
    )

    password = db.Column(
        db.String(255),
        nullable=False
    )

    name = db.Column(
        db.String(100),
        nullable=False
    )

    role = db.Column(
        db.String(20),
        default="user"
    )

    office = db.Column(
        db.String(100)
    )

    favorite_vehicles_json = db.Column(
        db.Text,
        default="[]"
    )
    
class Vehicle(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)

    # システム内で使用する車両ID
    vehicle_id = db.Column(db.String(50), nullable=False)

    # 車番・ナンバー
    plate_area = db.Column(db.String(50))
    plate_class = db.Column(db.String(50))
    plate_kana = db.Column(db.String(10))
    plate_number = db.Column(db.String(50))

    # 車両台帳情報
    chassis_number = db.Column(db.String(100))
    model_code = db.Column(db.String(100))
    first_registration_date = db.Column(db.String(20))
    manufacturer = db.Column(db.String(100))
    body_type = db.Column(db.String(100))

    gross_vehicle_weight = db.Column(db.Integer)
    max_payload = db.Column(db.Integer)

    # 既存項目
    type = db.Column(db.String(100))
    office = db.Column(db.String(100))
    inspection_expiry = db.Column(db.String(20))

    # 廃車・登録外になってもデータ自体は残す
    deleted = db.Column(
        db.Boolean,
        default=False
    )

class VehicleType(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(100), nullable=False)

class LicenseType(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(100), nullable=False)

class Driver(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)
    employee_id = db.Column(db.String(50), nullable=False)

    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(50))
    office = db.Column(db.String(100))
    safe_start_date = db.Column(db.String(20))

    vehicles_json = db.Column(db.Text, default="[]")
    licenses_json = db.Column(db.Text, default="[]")

class News(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text)
    files_json = db.Column(db.Text, default="[]")

    target_type = db.Column(db.String(50))
    target_value = db.Column(db.String(100))

    created_at = db.Column(db.String(20))


class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    target_user = db.Column(db.String(100), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text)
    link = db.Column(db.String(200))
    files_json = db.Column(db.Text, default="[]")

    read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.String(20))

class Office(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(100), nullable=False)

class DeliveryPlace(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(100), nullable=False)

class PatrolContentType(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(100), nullable=False)

class Manual(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    category = db.Column(db.String(100))
    filename = db.Column(db.String(200))

class VehiclePatrol(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)

    vehicle_id = db.Column(db.String(50))
    occurred_date = db.Column(db.String(20))
    category = db.Column(db.String(100))
    priority = db.Column(db.String(50))
    content = db.Column(db.Text)

    cause = db.Column(db.Text)
    temporary_action = db.Column(db.Text)
    repair_content = db.Column(db.Text)

    status = db.Column(db.String(50), default="未対応")

    repair_date = db.Column(db.String(20))
    repair_person = db.Column(db.String(100))
    repair_time = db.Column(db.String(50))
    parts = db.Column(db.Text)
    cost = db.Column(db.String(50))

class Checklist(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(db.String(50), nullable=False)

    name = db.Column(db.String(200), nullable=False)
    target = db.Column(db.String(100))

    frequency_value = db.Column(db.String(20))
    frequency_unit = db.Column(db.String(20))
    display_type = db.Column(db.String(20))
    print_portrait = db.Column(
        db.Boolean,
        default=False,
        nullable=False
    )

    items_json = db.Column(db.Text, default="[]")
    notify_users_json = db.Column(
        db.Text,
        default="[]"
    )

class ChecklistResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(
        db.String(50),
        nullable=False
    )

    checklist_id = db.Column(
        db.Integer,
        nullable=False
    )

    target_type = db.Column(db.String(50))
    target_user = db.Column(db.String(100))
    target_vehicle = db.Column(db.String(100))
    target_office = db.Column(db.String(100))

    checked_by = db.Column(db.String(100))
    checked_date = db.Column(db.String(30))

    status = db.Column(db.String(30))

    approved_by = db.Column(db.String(100))
    approved_date = db.Column(db.String(30))
    reject_reason = db.Column(db.Text)
    
    approvals_json = db.Column(
        db.Text,
        default="[]"
    )

    answers_json = db.Column(
        db.Text,
        default="[]"
    )

class VehicleChecklistResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(
        db.String(50),
        nullable=False
    )

    checklist_id = db.Column(
        db.Integer,
        nullable=False
    )

    vehicle_id = db.Column(
        db.String(50)
    )

    year = db.Column(db.String(10))
    month = db.Column(db.String(10))
    day = db.Column(db.String(10))

    checked_by = db.Column(db.String(100))
    checked_date = db.Column(db.String(30))

    status = db.Column(db.String(30))

    approved_by = db.Column(db.String(100))
    approved_date = db.Column(db.String(30))

    reject_reason = db.Column(db.Text)
    
    approvals_json = db.Column(
        db.Text,
        default="[]"
    )
    
    notify_users_json = db.Column(
        db.Text,
        default="[]"
    )

    answers_json = db.Column(
        db.Text,
        default="[]"
    )

class PatrolResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    company_code = db.Column(
        db.String(50),
        nullable=False
    )

    created_by_username = db.Column(
        db.String(100)
    )

    created_by_name = db.Column(
        db.String(100)
    )

    date = db.Column(
        db.String(20)
    )

    office = db.Column(
        db.String(100)
    )

    delivery_place = db.Column(
        db.String(100)
    )

    category = db.Column(
        db.String(50)
    )

    content_type = db.Column(
        db.String(50)
    )

    target_type = db.Column(
        db.String(50)
    )

    target_user = db.Column(
        db.String(100)
    )

    content = db.Column(
        db.Text
    )

    files_json = db.Column(
        db.Text,
        default="[]"
    )

    countermeasure = db.Column(
        db.Text
    )

    countermeasure_by = db.Column(
        db.String(100)
    )

    approval_status = db.Column(
        db.String(30)
    )

    reject_reason = db.Column(
        db.Text
    )

UPLOAD_FOLDER = "static/uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

S3_BUCKET_NAME = os.environ.get("S3_BUCKET_NAME", "")
AWS_REGION = os.environ.get("AWS_REGION", "ap-northeast-1")

s3_client = None

if S3_BUCKET_NAME:
    s3_client = boto3.client(
        "s3",
        region_name=AWS_REGION,
        aws_access_key_id=os.environ.get("AWS_ACCESS_KEY_ID"),
        aws_secret_access_key=os.environ.get("AWS_SECRET_ACCESS_KEY")
    )


def save_uploaded_file(file, folder=None):
    if not file or not file.filename:
        return ""

    original_filename = secure_filename(file.filename)
    extension = os.path.splitext(original_filename)[1].lower()
    filename = f"{uuid4().hex}{extension}"

    # Renderなど、S3設定がある環境
    if s3_client and S3_BUCKET_NAME:
        if folder == "static/manuals":
            s3_folder = "manuals"
        else:
            s3_folder = "uploads"

        object_key = f"{s3_folder}/{filename}"

        s3_client.upload_fileobj(
            file,
            S3_BUCKET_NAME,
            object_key,
            ExtraArgs={
                "ContentType": file.mimetype or "application/octet-stream"
            }
        )

        return filename

    # ローカル開発環境では従来どおり保存
    save_folder = folder or app.config["UPLOAD_FOLDER"]
    os.makedirs(save_folder, exist_ok=True)

    save_path = os.path.join(save_folder, filename)
    file.save(save_path)

    return filename

@app.route("/files/<folder>/<filename>")
def uploaded_file(folder, filename):
    if folder not in {"uploads", "manuals"}:
        return "Not found", 404

    # RenderなどS3利用環境
    if s3_client and S3_BUCKET_NAME:
        object_key = f"{folder}/{filename}"

        try:
            url = s3_client.generate_presigned_url(
                "get_object",
                Params={
                    "Bucket": S3_BUCKET_NAME,
                    "Key": object_key
                },
                ExpiresIn=3600
            )

            return redirect(url)

        except ClientError as e:
            print("S3取得エラー:", e)
            return "File not found", 404

    # ローカル環境
    if folder == "manuals":
        local_path = f"/static/manuals/{filename}"
    else:
        local_path = f"/static/uploads/{filename}"

    return redirect(local_path)

@app.route("/static/uploads/<path:filename>")
def s3_uploads_file(filename):
    if s3_client and S3_BUCKET_NAME:
        try:
            url = s3_client.generate_presigned_url(
                "get_object",
                Params={
                    "Bucket": S3_BUCKET_NAME,
                    "Key": f"uploads/{filename}"
                },
                ExpiresIn=3600
            )
            return redirect(url)

        except ClientError as e:
            print("S3取得エラー:", e)
            return "File not found", 404

    return app.send_static_file(f"uploads/{filename}")


@app.route("/static/manuals/<path:filename>")
def s3_manual_file(filename):
    if s3_client and S3_BUCKET_NAME:
        try:
            url = s3_client.generate_presigned_url(
                "get_object",
                Params={
                    "Bucket": S3_BUCKET_NAME,
                    "Key": f"manuals/{filename}"
                },
                ExpiresIn=3600
            )
            return redirect(url)

        except ClientError as e:
            print("S3取得エラー:", e)
            return "File not found", 404

    return app.send_static_file(f"manuals/{filename}")

PATROL_VIEW_TYPES = {"user", "delivery_place"}

@app.before_request
def require_login():
    if request.endpoint in {"login", "register", "static"} or request.endpoint is None:
        return None

    if not session.get("username"):
        return redirect("/login")

    return None


def require_itc():
    return session.get("role") == "itc"

def get_company(company_code):
    return Company.query.filter_by(
        company_code=company_code
    ).first()


def offices_for_current_company():
    query = Office.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        {
            "id": office.id,
            "index": office.id,
            "company_code": office.company_code,
            "name": office.name
        }
        for office in query.all()
    ]

def delivery_places_for_current_company():
    query = DeliveryPlace.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        {
            "id": place.id,
            "index": place.id,
            "company_code": place.company_code,
            "name": place.name
        }
        for place in query.all()
    ]

def patrol_content_types_for_current_company():
    query = PatrolContentType.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        {
            "id": item.id,
            "index": item.id,
            "company_code": item.company_code,
            "name": item.name
        }
        for item in query.all()
    ]

def manuals_for_current_company():
    query = Manual.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        {
            "id": manual.id,
            "index": manual.id,
            "company_code": manual.company_code,
            "title": manual.title,
            "category": manual.category,
            "filename": manual.filename,
        }
        for manual in query.all()
    ]

def vehicle_patrol_to_dict(patrol):
    return {
        "id": patrol.id,
        "index": patrol.id,
        "company_code": patrol.company_code,
        "vehicle_id": patrol.vehicle_id,
        "occurred_date": patrol.occurred_date,
        "category": patrol.category,
        "priority": patrol.priority,
        "content": patrol.content,
        "cause": patrol.cause,
        "temporary_action": patrol.temporary_action,
        "repair_content": patrol.repair_content,
        "status": patrol.status,
        "repair_date": patrol.repair_date,
        "repair_person": patrol.repair_person,
        "repair_time": patrol.repair_time,
        "parts": patrol.parts,
        "cost": patrol.cost,
    }


def vehicle_patrols_for_current_company():
    query = VehiclePatrol.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        vehicle_patrol_to_dict(patrol)
        for patrol in query.order_by(VehiclePatrol.id.desc()).all()
    ]

def checklist_to_dict(checklist):
    items = json.loads(checklist.items_json or "[]")

    return {
        "id": checklist.id,
        "index": checklist.id,
        "company_code": checklist.company_code,
        "name": checklist.name,
        "target": checklist.target,
        "frequency_value": checklist.frequency_value,
        "frequency_unit": checklist.frequency_unit,
        "display_type": checklist.display_type,
        "print_portrait": bool(checklist.print_portrait),
        "items": items,
        "score_enabled": any(
            item.get("score_enabled", False)
            for item in items
            if item.get("item_type") == "check"
        ),
    }


def checklists_for_current_company():
    query = Checklist.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        checklist_to_dict(checklist)
        for checklist in query.all()
    ]

def checklist_result_to_dict(result):
    return {
        "id": result.id,
        "index": result.id,
        "company_code": result.company_code,
        "checklist_id": result.checklist_id,
        "target_type": result.target_type,
        "target_user": result.target_user,
        "target_vehicle": result.target_vehicle,
        "target_office": result.target_office,
        "checked_by": result.checked_by,
        "checked_date": result.checked_date,
        "status": result.status,
        "approved_by": result.approved_by,
        "approved_date": result.approved_date,
        "reject_reason": result.reject_reason,
        "approvals": json.loads(result.approvals_json or "[]"),
        "answers": json.loads(result.answers_json or "[]"),
    }

def vehicle_checklist_result_to_dict(result):
    return {
        "id": result.id,
        "index": result.id,
        "company_code": result.company_code,
        "checklist_id": result.checklist_id,
        "vehicle_id": result.vehicle_id,
        "year": result.year,
        "month": result.month,
        "day": result.day,
        "checked_by": result.checked_by,
        "checked_date": result.checked_date,
        "status": result.status,
        "approved_by": result.approved_by,
        "approved_date": result.approved_date,
        "reject_reason": result.reject_reason,
        "approvals": json.loads(
            result.approvals_json or "[]"
        ),
        "notify_users": json.loads(
            result.notify_users_json or "[]"
        ),
        "answers": json.loads(result.answers_json or "[]"),
    }

def patrol_result_to_dict(result):
    return {
        "id": result.id,
        "index": result.id,
        "company_code": result.company_code,
        "created_by_username": result.created_by_username,
        "created_by_name": result.created_by_name,
        "date": result.date,
        "office": result.office,
        "delivery_place": result.delivery_place,
        "category": result.category,
        "content_type": result.content_type,
        "target_type": result.target_type,
        "target_user": result.target_user,
        "content": result.content,
        "files": json.loads(result.files_json or "[]"),
        "countermeasure": result.countermeasure,
        "approval_status": result.approval_status,
        "reject_reason": result.reject_reason,
    }


def patrol_results_for_current_company():
    query = PatrolResult.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        patrol_result_to_dict(result)
        for result in query.order_by(PatrolResult.id.desc()).all()
    ]

def add_notification(target_user, title, message, link="", files=None):
    if not target_user:
        return

    notification = Notification(
        target_user=target_user,
        title=title,
        message=message,
        link=link,
        files_json=json.dumps(files or [], ensure_ascii=False),
        read=False,
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M")
    )

    db.session.add(notification)
    db.session.commit()


def add_news(title, message, files=None, target_type="", target_value=""):
    news = News(
        title=title,
        message=message,
        files_json=json.dumps(files or [], ensure_ascii=False),
        target_type=target_type,
        target_value=target_value,
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M")
    )

    db.session.add(news)
    db.session.commit()


def notify_mentions(text, link=""):
    if not text:
        return

    user_query = User.query

    if session.get("role") != "itc":
        user_query = user_query.filter_by(
            company_code=session.get("company_code")
        )

    for user in user_query.all():
        name = user.name

        if not name:
            continue

        mention_text = "@" + name

        if mention_text in text:
            add_notification(
                name,
                "メンションされました",
                text,
                link
            )


def is_same_company_result(result):
    company_code = result.get("company_code")
    return not company_code or company_code == session.get("company_code")


def can_view_patrol_result(result):
    role = session.get("role")

    if role in ["itc", "admin"]:
        return is_same_company_result(result)

    if not is_same_company_result(result):
        return False

    if result.get("target_type") == "delivery_place":
        return True

    if result.get("created_by_username") == session.get("username"):
        return True

    return (
        result.get("target_type") == "user"
        and result.get("target_user") == session.get("name")
    )


def can_manage_patrol_result(result):
    return is_same_company_result(result)


def can_approve_patrol_result(result):
    return session.get("role") in ["admin", "itc"] and can_manage_patrol_result(result)


def can_view_checklist_result(result):
    role = session.get("role")
    name = session.get("name")

    if role in ["admin", "itc"]:
        return True

    if result.get("checked_by") == name:
        return True

    return result.get("target_type") == "user" and result.get("target_user") == name


def can_manage_checklist_result(result):
    role = session.get("role")

    if role in ["admin", "itc"]:
        return True

    return result.get("checked_by") == session.get("name")


def can_approve_checklist_result(result, approval=None):
    role = session.get("role")

    if role in ["admin", "itc"]:
        return True

    if not approval:
        return False

    return approval.get("allow_general", False)


def can_reject_checklist_result(result):
    return session.get("role") in ["admin", "itc"]

def vehicle_number(vehicle):
    return vehicle.get("number") or (
        f"{vehicle.get('plate_area', '')} "
        f"{vehicle.get('plate_class', '')} "
        f"{vehicle.get('plate_kana', '')} "
        f"{vehicle.get('plate_number', '')}"
    ).strip()


def delete_vehicle_related_data(vehicle):

    vehicle_id = vehicle.vehicle_id
    company_code = vehicle.company_code


    # 車両パトロール・修理履歴を削除
    VehiclePatrol.query.filter_by(
        company_code=company_code,
        vehicle_id=vehicle_id
    ).delete(synchronize_session=False)


    # 車両点検・チェックリスト履歴を削除
    VehicleChecklistResult.query.filter_by(
        company_code=company_code,
        vehicle_id=vehicle_id
    ).delete(synchronize_session=False)


    # 汎用チェックリストで
    # この車両を対象にした履歴を削除
    ChecklistResult.query.filter_by(
        company_code=company_code,
        target_vehicle=vehicle_id
    ).delete(synchronize_session=False)


def vehicles_with_numbers(include_inactive=False):

    query = Vehicle.query.filter_by(
        company_code=session.get("company_code")
    )

    if not include_inactive:
        query = query.filter_by(
            deleted=False
        )

    vehicles = []

    for vehicle in query.all():

        item = {
            "index": vehicle.id,
            "id": vehicle.id,
            "company_code": vehicle.company_code,
            "vehicle_id": vehicle.vehicle_id,
            "deleted": vehicle.deleted,

            "plate_area": vehicle.plate_area,
            "plate_class": vehicle.plate_class,
            "plate_kana": vehicle.plate_kana,
            "plate_number": vehicle.plate_number,

            "chassis_number": vehicle.chassis_number,
            "model_code": vehicle.model_code,
            "first_registration_date": vehicle.first_registration_date,
            "manufacturer": vehicle.manufacturer,
            "body_type": vehicle.body_type,

            "gross_vehicle_weight": vehicle.gross_vehicle_weight,
            "max_payload": vehicle.max_payload,

            "type": vehicle.type,
            "office": vehicle.office,
            "inspection_expiry": vehicle.inspection_expiry,
        }

        item["number"] = vehicle_number(item)

        vehicles.append(item)

    return vehicles

def vehicle_types_for_current_company():
    query = VehicleType.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        {
            "id": vehicle_type.id,
            "index": vehicle_type.id,
            "company_code": vehicle_type.company_code,
            "name": vehicle_type.name
        }
        for vehicle_type in query.all()
    ]

def license_types_for_current_company():
    query = LicenseType.query.filter_by(
        company_code=session.get("company_code")
    )
    
    return [
        {
            "id": license_type.id,
            "index": license_type.id,
            "company_code": license_type.company_code,
            "name": license_type.name
        }
        for license_type in query.all()
    ]

def driver_to_dict(driver):
    user = User.query.filter_by(
        company_code=driver.company_code,
        username=driver.employee_id
    ).first()

    return {
        "index": driver.id,
        "id": driver.id,
        "company_code": driver.company_code,
        "employee_id": driver.employee_id,
        "username": user.username if user else driver.employee_id,
        "name": driver.name,
        "role": driver.role,
        "office": driver.office,
        "safe_start_date": driver.safe_start_date,
        "vehicles": json.loads(driver.vehicles_json or "[]"),
        "licenses": json.loads(driver.licenses_json or "[]"),
    }


def drivers_for_current_company():
    query = Driver.query.filter_by(
        company_code=session.get("company_code")
    )

    return [
        driver_to_dict(driver)
        for driver in query.all()
    ]

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        company_code = request.form.get("company_code")
        username = request.form.get("username")
        password = request.form.get("password")

        user = User.query.filter_by(
            company_code=company_code,
            username=username
        ).first()

        if user and check_password_hash(user.password, password):

            company = get_company(user.company_code)

            if user.role != "itc":
                if not company or not company.active:
                    error = "この会社は現在利用停止中です。ITCへお問い合わせください。"
                else:
                    session.clear()
                    session["company_code"] = user.company_code
                    session["username"] = user.username
                    session["role"] = user.role
                    session["name"] = user.name
                    session["office"] = user.office
                    session["vehicles"] = json.loads(user.favorite_vehicles_json or "[]")

                    return redirect("/")
            else:
                session.clear()
                session["company_code"] = user.company_code
                session["username"] = user.username
                session["role"] = user.role
                session["name"] = user.name
                session["office"] = user.office
                session["vehicles"] = json.loads(user.favorite_vehicles_json or "[]")

                return redirect("/itc")

        else:
            error = "会社コード、ユーザーID、またはパスワードが違います。"

    return render_template("login.html", error=error)

@app.route("/itc/news/new", methods=["GET", "POST"])
def itc_new_news():
    if not require_itc():
        return redirect("/")

    if request.method == "POST":
        title = request.form.get("title")
        message = request.form.get("message")
        target_type = request.form.get("target_type")
        target_value = request.form.get("target_value", "").strip()

        file_names = []

        uploaded_files = request.files.getlist("files")

        for file in uploaded_files:
            filename = save_uploaded_file(file)

            if filename:
                file_names.append(filename)

        add_news(
            title,
            message,
            files=file_names,
            target_type=target_type,
            target_value=target_value
        )



        target_users = []

        user_query = User.query

        if target_type == "all":
            target_users = user_query.all()

        elif target_type == "admins":
            target_users = user_query.filter_by(
                role="admin"
            ).all()

        elif target_type == "company":
            target_users = user_query.filter_by(
                company_code=target_value
            ).all()

        elif target_type == "office":
            target_users = user_query.filter_by(
                office=target_value
            ).all()

        elif target_type == "user":
            target_users = user_query.filter_by(
                username=target_value
            ).all()

        for user in target_users:
            add_notification(
                user.name,
                title,
                message,
                "",
                files=file_names
            )
        notify_mentions(message, "/notifications")

        return redirect("/itc")

    return render_template(
        "itc_news_form.html",
        companies=Company.query.all(),
        offices=offices_for_current_company(),
        users=User.query.filter(User.role != "itc").all(),
        news=None,
        mode="new"
    )

@app.route("/itc/news/<int:index>/edit", methods=["GET", "POST"])
def itc_edit_news(index):
    if not require_itc():
        return redirect("/")

    news = News.query.get(index)

    if not news:
        return redirect("/itc")

    if request.method == "POST":
        news.title = request.form.get("title")
        news.message = request.form.get("message")
        news.target_type = request.form.get("target_type")
        news.target_value = request.form.get("target_value", "").strip()

        files = json.loads(news.files_json or "[]")

        uploaded_files = request.files.getlist("files")

        for file in uploaded_files:
            filename = save_uploaded_file(file)

            if filename:
                files.append(filename)

        news.files_json = json.dumps(files, ensure_ascii=False)

        db.session.commit()

        return redirect("/itc")

    news_dict = {
        "id": news.id,
        "index": news.id,
        "title": news.title,
        "message": news.message,
        "files": json.loads(news.files_json or "[]"),
        "target_type": news.target_type,
        "target_value": news.target_value,
        "created_at": news.created_at,
    }

    return render_template(
        "itc_news_form.html",
        news=news_dict,
        index=news.id,
        mode="edit",
        companies=Company.query.all(),
        offices=offices_for_current_company(),
        users=User.query.filter(User.role != "itc").all()
    )

@app.route("/itc/news/<int:index>/delete", methods=["POST"])
def itc_delete_news(index):
    if not require_itc():
        return redirect("/")

    news = News.query.get(index)

    if not news:
        return redirect("/itc")

    db.session.delete(news)
    db.session.commit()

    return redirect("/itc")

@app.context_processor
def inject_notification_count():
    user_name = session.get("name")

    unread_count = 0

    if user_name:
        unread_count = Notification.query.filter_by(
            target_user=user_name,
            read=False
        ).count()

    return {
        "unread_notification_count": unread_count
    }


@app.route("/register", methods=["GET", "POST"])
def register():
    error = None
    company_code_from_url = request.args.get("company_code", "")

    if request.method == "POST":
        company_code = request.form.get("company_code", "").strip()
        name = request.form.get("name", "").strip()
        employee_id = request.form.get("employee_id", "").strip()
        password = request.form.get("password", "").strip()
        office = request.form.get("office", "").strip()
        role = request.form.get("role", "user")

        company = get_company(company_code)

        if not company:
            error = "会社コードが存在しません。"

        elif not company.active:
            error = "この会社は現在利用停止中です。ITCへお問い合わせください。"

        elif role not in ["admin", "user"]:
            error = "ユーザー種別が不正です。"

        elif User.query.filter_by(
            company_code=company_code,
            username=employee_id
        ).first():
            error = "このログインIDはすでに使用されています。"

        else:
            user = User(
                company_code=company_code,
                username=employee_id,
                password=generate_password_hash(password),
                role=role,
                name=name,
                office=office,
                favorite_vehicles_json="[]"
            )

            driver = Driver(
                company_code=company_code,
                employee_id=employee_id,
                name=name,
                role=role,
                office=office,
                safe_start_date=datetime.now().strftime("%Y-%m-%d"),
                vehicles_json="[]",
                licenses_json="[]"
            )

            db.session.add(user)
            db.session.add(driver)
            db.session.commit()

            return redirect("/login")

    company_code = company_code_from_url

    offices = Office.query.filter_by(
        company_code=company_code
    ).all()

    return render_template(
        "register.html",
        error=error,
        company_code=company_code_from_url,
        offices=offices
    )

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/settings", methods=["GET", "POST"])
def settings():
    current_user = User.query.filter_by(
        company_code=session.get("company_code"),
        username=session.get("username")
    ).first()

    if not current_user:
        return redirect("/logout")

    error = None
    success = None

    if request.method == "POST":
        current_password = request.form.get("current_password")
        new_password = request.form.get("new_password")
        new_password_confirm = request.form.get("new_password_confirm")

        if not check_password_hash(current_user.password, current_password):
            error = "現在のパスワードが違います。"

        elif not new_password:
            error = "新しいパスワードを入力してください。"

        elif new_password != new_password_confirm:
            error = "新しいパスワードが一致しません。"

        else:
            current_user.password = generate_password_hash(new_password)
            db.session.commit()
            success = "パスワードを変更しました。"

    return render_template(
        "settings.html",
        user=current_user,
        error=error,
        success=success
    )

@app.route("/api/news-targets")
def news_targets():
    target_type = request.args.get("type", "")
    keyword = request.args.get("q", "").strip()

    results = []

    if target_type == "company":
        query = Company.query

        if keyword:
            keyword_like = f"%{keyword}%"
            query = query.filter(
                db.or_(
                    Company.company_name.ilike(keyword_like),
                    Company.company_code.ilike(keyword_like)
                )
            )

        for company in query.order_by(
            Company.company_name.asc()
        ).limit(10).all():
            results.append({
                "id": company.company_code,
                "name": company.company_name,
                "sub": "会社コード：" + company.company_code
            })

    elif target_type == "office":
        query = Office.query

        if keyword:
            keyword_like = f"%{keyword}%"
            query = query.filter(
                db.or_(
                    Office.name.ilike(keyword_like),
                    Office.company_code.ilike(keyword_like)
                )
            )

        for office in query.order_by(
            Office.name.asc()
        ).limit(10).all():
            results.append({
                "id": office.name or "",
                "name": office.name or "",
                "sub": "営業所 / " + (office.company_code or "")
            })

    elif target_type == "user":
        query = User.query.filter(
            User.role != "itc"
        )

        if keyword:
            keyword_like = f"%{keyword}%"
            query = query.filter(
                db.or_(
                    User.name.ilike(keyword_like),
                    User.username.ilike(keyword_like),
                    User.office.ilike(keyword_like),
                    User.company_code.ilike(keyword_like)
                )
            )

        for user in query.order_by(
            User.name.asc()
        ).limit(10).all():
            results.append({
                "id": user.username,
                "name": user.name or "",
                "sub": (user.office or "") + " / " + (user.company_code or "")
            })

    return {"results": results}

@app.route("/api/mention-users")
def mention_users():
    keyword = request.args.get("q", "").strip()

    users = []

    user_query = User.query

    if session.get("role") != "itc":
        user_query = user_query.filter_by(
            company_code=session.get("company_code")
        )

    for user in user_query.all():
        name = user.name or ""

        if keyword and keyword not in name:
            continue

        users.append({
            "name": name,
            "username": user.username,
            "office": user.office or ""
        })

    for driver in drivers_for_current_company():
        name = driver.get("name", "")

        if keyword and keyword not in name:
            continue

        if not any(u["name"] == name for u in users):
            users.append({
                "name": name,
                "username": driver.get("employee_id"),
                "office": driver.get("office", "")
            })

    return {"users": users[:10]}

@app.route("/api/vehicles")
def search_vehicles():

    keyword = request.args.get("q", "").strip()

    company_code = request.args.get(
        "company_code",
        session.get("company_code")
    )

    if session.get("role") != "itc":
        company_code = session.get("company_code")

    query = Vehicle.query.filter_by(
        company_code=company_code,
        deleted=False
    )
    if keyword:

        keyword_like = f"%{keyword}%"

        query = query.filter(
            db.or_(
                Vehicle.vehicle_id.ilike(keyword_like),
                Vehicle.plate_area.ilike(keyword_like),
                Vehicle.plate_class.ilike(keyword_like),
                Vehicle.plate_kana.ilike(keyword_like),
                Vehicle.plate_number.ilike(keyword_like),
                Vehicle.chassis_number.ilike(keyword_like),
                Vehicle.model_code.ilike(keyword_like),
                Vehicle.manufacturer.ilike(keyword_like),
            )
        )

    vehicles = (
        query
        .order_by(Vehicle.id.asc())
        .limit(20)
        .all()
    )

    results = []

    for vehicle in vehicles:

        number = " ".join(
            value
            for value in [
                vehicle.plate_area or "",
                vehicle.plate_class or "",
                vehicle.plate_kana or "",
                vehicle.plate_number or "",
            ]
            if value
        )

        results.append({
            "id": vehicle.vehicle_id,
            "vehicle_id": vehicle.vehicle_id,
            "number": number,
            "chassis_number": vehicle.chassis_number or "",
            "manufacturer": vehicle.manufacturer or "",
            "model_code": vehicle.model_code or "",
        })

    return {
        "results": results
    }
    
@app.route("/")
def dashboard():
    today = datetime.now().date()
    user_name = session.get("name")

    company_code = session.get("company_code")

    # 自分の無事故無違反日数
    my_driver = Driver.query.filter_by(
        company_code=company_code,
        name=user_name
    ).first()

    my_safe_days = 0

    if my_driver and my_driver.safe_start_date:
        start_date = datetime.strptime(
            my_driver.safe_start_date,
            "%Y-%m-%d"
        ).date()
    
        my_safe_days = (today - start_date).days

    # 社内ランキング
    ranking_records = Driver.query.filter(
        Driver.company_code == company_code,
        Driver.safe_start_date.isnot(None),
        Driver.safe_start_date != ""
    ).order_by(
        Driver.safe_start_date.asc()
    ).limit(10).all()

    ranking = []

    for driver in ranking_records:
        start_date = datetime.strptime(
            driver.safe_start_date,
            "%Y-%m-%d"
        ).date()

        ranking.append({
            "id": driver.id,
            "index": driver.id,
            "employee_id": driver.employee_id,
            "name": driver.name,
            "role": driver.role,
            "office": driver.office,
            "safe_start_date": driver.safe_start_date,
            "safe_days": (today - start_date).days,
        })

    # 自分のGood件数
    my_good_count = PatrolResult.query.filter_by(
        company_code=session.get("company_code"),
        category="Good",
        target_user=user_name
    ).count()

    # 自分に対する未対応指摘
    my_pending_pointouts = []

    pending_records = PatrolResult.query.filter_by(
        company_code=session.get("company_code"),
        target_type="user",
        target_user=user_name
    ).filter(
        PatrolResult.category != "Good",
        PatrolResult.approval_status != "承認済み"
    ).order_by(
        PatrolResult.id.desc()
    ).all()

    for record in pending_records:
        my_pending_pointouts.append(
            patrol_result_to_dict(record)
        )

    # 車検が近い車両
    inspection_alerts = []

    inspection_limit_date = (
        today + timedelta(days=90)
    ).strftime("%Y-%m-%d")


    inspection_vehicle_records = Vehicle.query.filter(
        Vehicle.company_code == session.get("company_code"),
        Vehicle.deleted == False,
        Vehicle.inspection_expiry.isnot(None),
        Vehicle.inspection_expiry != "",
        Vehicle.inspection_expiry <= inspection_limit_date
    ).order_by(
        Vehicle.inspection_expiry.asc()
    ).all()


    for vehicle in inspection_vehicle_records:

        try:
            expiry_date = datetime.strptime(
                vehicle.inspection_expiry,
                "%Y-%m-%d"
            ).date()
        except ValueError:
            continue


        remaining_days = (
            expiry_date - today
        ).days


        item = {
            "vehicle_id": vehicle.vehicle_id,
            "number": " ".join(
                value
                for value in [
                    vehicle.plate_area or "",
                    vehicle.plate_class or "",
                    vehicle.plate_kana or "",
                    vehicle.plate_number or "",
                ]
                if value
            ),
            "inspection_expiry": vehicle.inspection_expiry,
            "remaining_days": remaining_days,
        }

        inspection_alerts.append(item)

    setup_tasks = []

    if session.get("role") == "admin":
        company_code = session.get("company_code")

        if Office.query.filter_by(company_code=company_code).count() == 0:
            setup_tasks.append({
                "name": "営業所マスタ",
                "description": "営業所を登録してください。",
                "url": "/master/offices"
            })

        if VehicleType.query.filter_by(company_code=company_code).count() == 0:
            setup_tasks.append({
                "name": "車種マスタ",
                "description": "車両登録で使用する車種を登録してください。",
                "url": "/master/vehicle-types"
            })

        if LicenseType.query.filter_by(company_code=company_code).count() == 0:
            setup_tasks.append({
                "name": "免許種別マスタ",
                "description": "ドライバー登録で使用する免許種別を登録してください。",
                "url": "/master/license-types"
            })

        if DeliveryPlace.query.filter_by(company_code=company_code).count() == 0:
            setup_tasks.append({
                "name": "納入先マスタ",
                "description": "安全パトロールで使用する納入先を登録してください。",
                "url": "/master/delivery-places"
            })

        if Vehicle.query.filter_by(
            company_code=company_code,
            deleted=False
        ).count() == 0:
            setup_tasks.append({
                "name": "車両マスタ",
                "description": "車両管理・点検で使用する車両を登録してください。",
                "url": "/master/vehicles"
            })

        if Checklist.query.filter_by(company_code=company_code).count() == 0:
            setup_tasks.append({
                "name": "チェックリストマスタ",
                "description": "安全管理・車両管理で使用するチェックリストを登録してください。",
                "url": "/master/checklists"
            })

    return render_template(
        "index.html",
        my_safe_days=my_safe_days,
        ranking=ranking,
        my_good_count=my_good_count,
        my_pending_pointouts=my_pending_pointouts,
        inspection_alerts=inspection_alerts,
        setup_tasks=setup_tasks
    )

@app.route("/notifications")
def notifications():
    user_name = session.get("name")

    items = []

    notification_records = Notification.query.filter_by(
        target_user=user_name
    ).order_by(Notification.id.desc()).all()

    for notification in notification_records:
        items.append({
            "id": notification.id,
            "index": notification.id,
            "target_user": notification.target_user,
            "title": notification.title,
            "message": notification.message,
            "link": notification.link,
            "files": json.loads(notification.files_json or "[]"),
            "read": notification.read,
            "created_at": notification.created_at,
        })

    return render_template(
        "notifications.html",
        notifications=items
    )

@app.route("/notifications/<int:index>")
def notification_detail(index):
    notification = Notification.query.get(index)

    if not notification:
        return redirect("/notifications")

    if notification.target_user != session.get("name"):
        return redirect("/notifications")

    notification.read = True
    db.session.commit()

    notification_dict = {
        "id": notification.id,
        "index": notification.id,
        "target_user": notification.target_user,
        "title": notification.title,
        "message": notification.message,
        "link": notification.link,
        "files": json.loads(notification.files_json or "[]"),
        "read": notification.read,
        "created_at": notification.created_at,
    }

    return render_template(
        "notification_detail.html",
        notification=notification_dict,
        index=notification.id
    )

@app.route("/notifications/<int:index>/delete", methods=["POST"])
def delete_notification(index):
    notification = Notification.query.get(index)

    if not notification:
        return redirect("/notifications")

    if notification.target_user != session.get("name"):
        return redirect("/notifications")

    db.session.delete(notification)
    db.session.commit()

    return redirect("/notifications")

@app.route("/itc")
def itc_dashboard():
    if not require_itc():
        return redirect("/")

    company_summaries = []

    companies = Company.query.all()

    for company in companies:
        company_code = company.company_code

        vehicle_count = Vehicle.query.filter_by(
            company_code=company_code,
            deleted=False
        ).count()

        item = {
            "id": company.id,
            "company_code": company.company_code,
            "company_name": company.company_name,
            "vehicle_limit": company.vehicle_limit,
            "active": company.active,
            "vehicle_count": vehicle_count,
            "remaining_vehicles": company.vehicle_limit - vehicle_count,
        }

        company_summaries.append(item)

    news_items = []

    for news in News.query.order_by(News.id.desc()).all():
        news_items.append({
            "id": news.id,
            "index": news.id,
            "title": news.title,
            "message": news.message,
            "files": json.loads(news.files_json or "[]"),
            "target_type": news.target_type,
            "target_value": news.target_value,
            "created_at": news.created_at,
        })

    return render_template(
        "itc_dashboard.html",
        companies=company_summaries,
        news=news_items
    )

@app.route("/itc/companies/new", methods=["GET", "POST"])
def itc_new_company():
    if not require_itc():
        return redirect("/")

    if request.method == "POST":
        company = Company(
            company_code=request.form.get("company_code"),
            company_name=request.form.get("company_name"),
            vehicle_limit=int(request.form.get("vehicle_limit") or 0),
            active=True,
        )

        db.session.add(company)
        db.session.commit()

        return redirect("/itc")

    return render_template(
        "itc_company_form.html",
        company=None,
        index=None,
        mode="new"
    )

@app.route("/itc/companies/<int:index>/edit", methods=["GET", "POST"])
def itc_edit_company(index):
    if not require_itc():
        return redirect("/")

    company = Company.query.get(index)

    if not company:
        return redirect("/itc")

    if request.method == "POST":
        company.company_code = request.form.get("company_code")
        company.company_name = request.form.get("company_name")
        company.vehicle_limit = int(request.form.get("vehicle_limit") or 0)
        company.active = request.form.get("active") == "1"

        db.session.commit()

        return redirect("/itc")

    company_dict = {
        "company_code": company.company_code,
        "company_name": company.company_name,
        "vehicle_limit": company.vehicle_limit,
        "active": company.active,
    }

    return render_template(
        "itc_company_form.html",
        company=company_dict,
        index=company.id,
        mode="edit"
    )

@app.route("/safety")
def safety():
    return render_template("safety.html")

@app.route("/pointouts")
def pointouts():
    role = session.get("role")

    view_type = request.args.get("type", "user")
    if view_type not in PATROL_VIEW_TYPES:
        view_type = "user"

    keyword = request.args.get("keyword", "").strip()

    query = PatrolResult.query.filter(
        PatrolResult.company_code == session.get("company_code"),
        PatrolResult.target_type == view_type
    )

    if keyword:
        keyword_like = f"%{keyword}%"

        if view_type == "user":
            query = query.filter(
                PatrolResult.target_user.ilike(keyword_like)
            )
        elif view_type == "delivery_place":
            query = query.filter(
                PatrolResult.delivery_place.ilike(keyword_like)
            )

    result_records = query.order_by(
        PatrolResult.id.desc()
    ).all()

    visible_results = []

    for result_record in result_records:
        result = patrol_result_to_dict(result_record)

        if not can_view_patrol_result(result):
            continue

        result["can_manage"] = can_manage_patrol_result(result)
        visible_results.append(result)

    driver_options = Driver.query.filter(
        Driver.company_code == session.get("company_code")
    ).order_by(
        Driver.name.asc()
    ).all()

    return render_template(
        "pointouts.html",
        patrol_results=visible_results,
        view_type=view_type,
        keyword=keyword,
        role=role,
        show_target_user=view_type == "user",
        drivers=driver_options,
        delivery_places=delivery_places_for_current_company(),
    )


@app.route("/pointouts/new", methods=["GET", "POST"])
def new_pointout():
    if request.method == "POST":
        uploaded_files = request.files.getlist("files")

        file_names = []

        for file in uploaded_files:
            filename = save_uploaded_file(file)

            if filename:
                file_names.append(filename)

        target_type = request.form.get("target_type")
        target_office = session.get("office")

        if target_type == "user":
            target_driver = Driver.query.filter_by(
                company_code=session.get("company_code"),
                name=request.form.get("target_user")
            ).first()

            if target_driver:
                target_office = target_driver.office

        if target_type not in PATROL_VIEW_TYPES:
            target_type = "user"

        result = PatrolResult(
            company_code=session.get("company_code"),
            created_by_username=session.get("username"),
            created_by_name=session.get("name"),
            date=request.form.get("date"),
            office=target_office,
            category=request.form.get("category"),
            content_type=request.form.get("content_type"),
            target_type=target_type,
            target_user=request.form.get("target_user") if target_type == "user" else "",
            delivery_place=request.form.get("delivery_place") if target_type == "delivery_place" else "",
            content=request.form.get("content"),
            files_json=json.dumps(file_names, ensure_ascii=False),
            countermeasure="",
            approval_status="未対応",
            reject_reason=""
        )

        db.session.add(result)
        db.session.commit()

        add_notification(
            result.target_user,
            "安全パトロール確認依頼",
            "あなたに確認が必要な安全パトロールがあります。",
            f"/pointouts/{result.id}"
        )

        notify_mentions(
            result.content,
            f"/pointouts/{result.id}"
        )

        if target_type == "delivery_place":
            return redirect("/pointouts?type=delivery_place")

        return redirect("/pointouts?type=user")

    return render_template(
        "new_pointout.html",
        drivers=drivers_for_current_company(),
        delivery_places=delivery_places_for_current_company(),
        manuals=manuals_for_current_company(),
        content_types=patrol_content_types_for_current_company(),
    )

@app.route("/pointouts/<int:index>")
def pointout_detail(index):
    result_record = PatrolResult.query.get(index)

    if not result_record:
        return redirect("/pointouts")

    result = patrol_result_to_dict(result_record)

    if not can_view_patrol_result(result):
        return redirect("/pointouts")

    return render_template(
        "pointout_detail.html",
        result=result,
        index=result_record.id,
        manuals=manuals_for_current_company(),
        can_manage=can_manage_patrol_result(result),
        can_approve=can_approve_patrol_result(result),
    )

@app.route("/pointouts/<int:index>/edit", methods=["GET", "POST"])
def edit_pointout(index):

    result_record = PatrolResult.query.get(index)

    if not result_record:
        return redirect("/pointouts")

    result = patrol_result_to_dict(result_record)

    if not can_manage_patrol_result(result):
        return redirect(f"/pointouts/{index}")

    if request.method == "POST":

        result_record.date = request.form.get("date")
        result_record.category = request.form.get("category")
        result_record.content_type = request.form.get("content_type")
        result_record.content = request.form.get("content")

        if result_record.target_type == "user":
            result_record.target_user = request.form.get("target_user")

            target_driver = Driver.query.filter_by(
                company_code=session.get("company_code"),
                name=result_record.target_user
            ).first()

            if target_driver:
                result_record.office = target_driver.office

        if result_record.target_type == "delivery_place":
            result_record.delivery_place = request.form.get("delivery_place")

        files = json.loads(result_record.files_json or "[]")

        delete_files = request.form.getlist("delete_files")

        for delete_file in delete_files:
            if delete_file in files:
                files.remove(delete_file)

                file_path = os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    delete_file
                )

                if os.path.exists(file_path):
                    os.remove(file_path)

        uploaded_files = request.files.getlist("files")

        for file in uploaded_files:
            filename = save_uploaded_file(file)

            if filename:
                files.append(filename)

        result_record.files_json = json.dumps(
            files,
            ensure_ascii=False
        )

        db.session.commit()

        return redirect(f"/pointouts/{result_record.id}")

    return render_template(
        "edit_pointout.html",
        result=result,
        index=result_record.id,
        drivers=drivers_for_current_company(),
        delivery_places=delivery_places_for_current_company(),
        manuals=manuals_for_current_company(),
    )

@app.route("/pointouts/<int:index>/countermeasure/new")
def new_countermeasure(index):
    result_record = PatrolResult.query.get(index)

    if not result_record:
        return redirect("/pointouts")

    result = patrol_result_to_dict(result_record)

    if not can_manage_patrol_result(result):
        return redirect(f"/pointouts/{index}")

    return render_template(
        "countermeasure.html",
        result=result,
        index=result_record.id
    )


@app.route("/pointouts/<int:index>/countermeasure", methods=["POST"])
def register_countermeasure(index):
    result_record = PatrolResult.query.get(index)

    if not result_record:
        return redirect("/pointouts")

    result = patrol_result_to_dict(result_record)

    if not can_manage_patrol_result(result):
        return redirect(f"/pointouts/{index}")

    result_record.countermeasure = request.form.get("countermeasure")
    result_record.countermeasure_by = session.get("name")
    result_record.approval_status = "承認待ち"
    result_record.reject_reason = ""

    db.session.commit()

    return redirect(f"/pointouts/{result_record.id}")


@app.route("/pointouts/<int:index>/approve", methods=["POST"])
def approve_countermeasure(index):
    result_record = PatrolResult.query.get(index)

    if not result_record:
        return redirect("/pointouts")

    result = patrol_result_to_dict(result_record)

    if not can_approve_patrol_result(result):
        return redirect(f"/pointouts/{index}")

    result_record.approval_status = "承認済み"
    result_record.reject_reason = ""

    db.session.commit()

    return redirect(f"/pointouts/{result_record.id}")


@app.route("/pointouts/<int:index>/reject", methods=["POST"])
def reject_countermeasure(index):
    result_record = PatrolResult.query.get(index)

    if not result_record:
        return redirect("/pointouts")

    result = patrol_result_to_dict(result_record)

    if not can_approve_patrol_result(result):
        return redirect(f"/pointouts/{index}")

    reject_reason = request.form.get("reject_reason", "")

    result_record.approval_status = "差し戻し"
    result_record.reject_reason = reject_reason

    notify_users = set()

    if result_record.created_by_name:
        notify_users.add(result_record.created_by_name)

    if result_record.target_user:
        notify_users.add(result_record.target_user)

    if result_record.countermeasure_by:
        notify_users.add(result_record.countermeasure_by)

    for target_user in notify_users:
        add_notification(
            target_user,
            "安全パトロールが差し戻されました",
            reject_reason or "安全パトロールが差し戻されました。",
            f"/pointouts/{result_record.id}"
        )

    db.session.commit()

    return redirect(f"/pointouts/{result_record.id}")


@app.route("/pointouts/<int:index>/delete", methods=["POST"])
def delete_pointout(index):
    result_record = PatrolResult.query.get(index)

    if not result_record:
        return redirect("/pointouts")

    result = patrol_result_to_dict(result_record)

    if not can_manage_patrol_result(result):
        return redirect(f"/pointouts/{index}")

    view_type = result.get("target_type", "user")

    db.session.delete(result_record)
    db.session.commit()

    return redirect(f"/pointouts?type={view_type}")

@app.route("/vehicle-patrols/new", methods=["GET", "POST"])
def new_vehicle_patrol():

    if request.method == "POST":

        patrol = VehiclePatrol(
            company_code=session.get("company_code"),
            vehicle_id=request.form.get("vehicle_id"),
            occurred_date=request.form.get("occurred_date"),
            category=request.form.get("category"),
            priority=request.form.get("priority"),
            content=request.form.get("content"),
            cause=request.form.get("cause"),
            temporary_action=request.form.get("temporary_action"),
            repair_content=request.form.get("repair_content"),
            status=request.form.get("status", "未対応"),
            repair_date=request.form.get("repair_date"),
            repair_person=request.form.get("repair_person"),
            repair_time=request.form.get("repair_time"),
            parts=request.form.get("parts"),
            cost=request.form.get("cost")
        )

        db.session.add(patrol)
        db.session.commit()

        return redirect("/vehicle-patrols")

    return render_template(
        "vehicle_patrol_form.html",
        patrol=None,
        selected_vehicle=None,
        mode="new"
    )


@app.route("/vehicle-favorites/add", methods=["POST"])
def add_vehicle_favorite():
    vehicle_id = request.form.get("vehicle_id", "").strip()

    if not vehicle_id:
        return redirect("/vehicle-patrols")

    vehicle = Vehicle.query.filter_by(
        company_code=session.get("company_code"),
        vehicle_id=vehicle_id,
        deleted=False
    ).first()

    if not vehicle:
        return redirect("/vehicle-patrols")

    current_user = User.query.filter_by(
        company_code=session.get("company_code"),
        username=session.get("username")
    ).first()

    if not current_user:
        return redirect("/logout")

    favorite_vehicles = json.loads(
        current_user.favorite_vehicles_json or "[]"
    )

    if vehicle_id not in favorite_vehicles:
        favorite_vehicles.append(vehicle_id)

    current_user.favorite_vehicles_json = json.dumps(
        favorite_vehicles,
        ensure_ascii=False
    )

    db.session.commit()

    session["vehicles"] = favorite_vehicles

    return redirect("/vehicle-patrols")

@app.route("/vehicle-favorites/remove/<vehicle_id>", methods=["POST"])
def remove_vehicle_favorite(vehicle_id):
    current_user = User.query.filter_by(
        company_code=session.get("company_code"),
        username=session.get("username")
    ).first()

    if not current_user:
        return redirect("/logout")

    favorite_vehicles = json.loads(
        current_user.favorite_vehicles_json or "[]"
    )

    if vehicle_id in favorite_vehicles:
        favorite_vehicles.remove(vehicle_id)

    current_user.favorite_vehicles_json = json.dumps(
        favorite_vehicles,
        ensure_ascii=False
    )

    db.session.commit()

    session["vehicles"] = favorite_vehicles

    return redirect("/vehicle-patrols")

@app.route("/vehicle-patrols")
def vehicle_patrols():

    keyword = request.args.get("keyword", "").strip()
    status_filter = request.args.get("status", "active")

    favorite_vehicles = session.get("vehicles", [])

    favorite_patrols = []
    other_patrols = []

    query = VehiclePatrol.query.filter(
        VehiclePatrol.company_code == session.get("company_code")
    )

    if status_filter == "active":
        query = query.filter(
            VehiclePatrol.status != "修理完了"
        )
    elif status_filter:
        query = query.filter(
            VehiclePatrol.status == status_filter
        )

    if keyword:
        keyword_like = f"%{keyword}%"

        query = query.filter(
            db.or_(
                VehiclePatrol.vehicle_id.ilike(keyword_like),
                VehiclePatrol.content.ilike(keyword_like),
                VehiclePatrol.repair_person.ilike(keyword_like)
            )
        )

    patrol_records = query.order_by(
        VehiclePatrol.id.desc()
    ).all()

    for patrol_record in patrol_records:
        patrol = vehicle_patrol_to_dict(patrol_record)

        if patrol.get("vehicle_id") in favorite_vehicles:
            favorite_patrols.append(patrol)
        else:
            other_patrols.append(patrol)

    return render_template(
        "vehicle_patrols.html",
        favorite_patrols=favorite_patrols,
        other_patrols=other_patrols,
        favorite_vehicles=favorite_vehicles,
        keyword=keyword,
        status_filter=status_filter
    )

@app.route("/vehicle-patrols/<int:index>")
def vehicle_patrol_detail(index):

    patrol = VehiclePatrol.query.get(index)

    if not patrol:
        return redirect("/vehicle-patrols")

    if session.get("role") != "itc":
        if patrol.company_code != session.get("company_code"):
            return redirect("/vehicle-patrols")

    return render_template(
        "vehicle_patrol_detail.html",
        patrol=vehicle_patrol_to_dict(patrol),
        index=patrol.id
    )

@app.route("/vehicle-patrols/<int:index>/edit", methods=["GET", "POST"])
def edit_vehicle_patrol(index):

    patrol = VehiclePatrol.query.get(index)

    if not patrol:
        return redirect("/vehicle-patrols")

    if session.get("role") != "itc":
        if patrol.company_code != session.get("company_code"):
            return redirect("/vehicle-patrols")

    if request.method == "POST":

        patrol.vehicle_id = request.form.get("vehicle_id")
        patrol.occurred_date = request.form.get("occurred_date")
        patrol.category = request.form.get("category")
        patrol.priority = request.form.get("priority")
        patrol.content = request.form.get("content")

        patrol.cause = request.form.get("cause")
        patrol.temporary_action = request.form.get("temporary_action")
        patrol.repair_content = request.form.get("repair_content")

        patrol.repair_date = request.form.get("repair_date")
        patrol.repair_person = request.form.get("repair_person")
        patrol.repair_time = request.form.get("repair_time")
        patrol.parts = request.form.get("parts")
        patrol.cost = request.form.get("cost")
        patrol.status = request.form.get("status", "未対応")

        db.session.commit()

        return redirect(f"/vehicle-patrols/{patrol.id}")

    selected_vehicle = None

    vehicle = Vehicle.query.filter_by(
        company_code=patrol.company_code,
        vehicle_id=patrol.vehicle_id
    ).first()

    if vehicle:

        number = " ".join(
            value
            for value in [
                vehicle.plate_area or "",
                vehicle.plate_class or "",
                vehicle.plate_kana or "",
                vehicle.plate_number or "",
            ]
            if value
        )

        selected_vehicle = {
            "vehicle_id": vehicle.vehicle_id,
            "number": number,
            "manufacturer": vehicle.manufacturer or "",
            "model_code": vehicle.model_code or "",
        }


    return render_template(
        "vehicle_patrol_form.html",
        patrol=vehicle_patrol_to_dict(patrol),
        index=patrol.id,
        selected_vehicle=selected_vehicle,
        mode="edit"
    )

@app.route("/vehicle-patrols/<int:index>/delete", methods=["POST"])
def delete_vehicle_patrol(index):

    patrol = VehiclePatrol.query.get(index)

    if not patrol:
        return redirect("/vehicle-patrols")

    if session.get("role") != "itc":
        if patrol.company_code != session.get("company_code"):
            return redirect("/vehicle-patrols")

    db.session.delete(patrol)
    db.session.commit()

    return redirect("/vehicle-patrols")

@app.route("/master/vehicle-types")
def vehicle_type_master():
    return render_template(
        "vehicle_type_master.html",
        vehicle_types=vehicle_types_for_current_company()
    )

@app.route("/master/license-types")
def license_type_master():
    return render_template(
        "license_type_master.html",
        license_types=license_types_for_current_company()
    )


@app.route("/master/license-types/new", methods=["GET", "POST"])
def new_license_type():
    if request.method == "POST":
        name = request.form.get("name", "").strip()

        if LicenseType.query.filter_by(
            company_code=session.get("company_code"),
            name=name
        ).first():
            return "この免許種別はすでに登録されています。"

        license_type = LicenseType(
            company_code=session.get("company_code"),
            name=name
        )

        db.session.add(license_type)
        db.session.commit()

        return redirect("/master/license-types")

    return render_template(
        "license_type_form.html",
        license_type=None,
        mode="new"
    )


@app.route("/master/license-types/<int:index>/edit", methods=["GET", "POST"])
def edit_license_type(index):
    license_type = LicenseType.query.get(index)

    if not license_type:
        return redirect("/master/license-types")

    if session.get("role") != "itc":
        if license_type.company_code != session.get("company_code"):
            return redirect("/master/license-types")

    if request.method == "POST":
        license_type.name = request.form.get("name")

        db.session.commit()

        return redirect("/master/license-types")

    return render_template(
        "license_type_form.html",
        license_type={
            "id": license_type.id,
            "name": license_type.name
        },
        mode="edit"
    )


@app.route("/master/license-types/<int:index>/delete", methods=["POST"])
def delete_license_type(index):
    license_type = LicenseType.query.get(index)

    if not license_type:
        return redirect("/master/license-types")

    if session.get("role") != "itc":
        if license_type.company_code != session.get("company_code"):
            return redirect("/master/license-types")
        
    for driver in Driver.query.filter_by(
        company_code=license_type.company_code
    ).all():
        licenses = json.loads(driver.licenses_json or "[]")

        licenses = [
            license
            for license in licenses
            if license.get("type") != license_type.name
        ]

        driver.licenses_json = json.dumps(
            licenses,
            ensure_ascii=False
        )

    db.session.delete(license_type)
    db.session.commit()

    return redirect("/master/license-types")

@app.route("/master/vehicle-types/new", methods=["GET", "POST"])
def new_vehicle_type():
    if request.method == "POST":
        name = request.form.get("name", "").strip()

        if VehicleType.query.filter_by(
            company_code=session.get("company_code"),
            name=name
        ).first():
            return "この車種はすでに登録されています。"

        vehicle_type = VehicleType(
            company_code=session.get("company_code"),
            name=name
        )

        db.session.add(vehicle_type)
        db.session.commit()

        return redirect("/master/vehicle-types")

    return render_template(
        "vehicle_type_form.html",
        vehicle_type=None,
        mode="new"
    )


@app.route("/master/vehicle-types/<int:index>/edit", methods=["GET", "POST"])
def edit_vehicle_type(index):
    vehicle_type = VehicleType.query.get(index)

    if not vehicle_type:
        return redirect("/master/vehicle-types")

    if session.get("role") != "itc":
        if vehicle_type.company_code != session.get("company_code"):
            return redirect("/master/vehicle-types")

    if request.method == "POST":
        vehicle_type.name = request.form.get("name")
        db.session.commit()

        return redirect("/master/vehicle-types")

    return render_template(
        "vehicle_type_form.html",
        vehicle_type={
            "id": vehicle_type.id,
            "name": vehicle_type.name
        },
        mode="edit"
    )


@app.route("/master/vehicle-types/<int:index>/delete", methods=["POST"])
def delete_vehicle_type(index):
    vehicle_type = VehicleType.query.get(index)

    if not vehicle_type:
        return redirect("/master/vehicle-types")

    if session.get("role") != "itc":
        if vehicle_type.company_code != session.get("company_code"):
            return redirect("/master/vehicle-types")

    Vehicle.query.filter_by(
        company_code=vehicle_type.company_code,
        type=vehicle_type.name,
        deleted=False
    ).update(
        {"type": ""},
        synchronize_session=False
    )

    db.session.delete(vehicle_type)
    db.session.commit()

    return redirect("/master/vehicle-types")

@app.route("/vehicle")
def vehicle():
    return render_template("vehicle.html")


@app.route("/analysis")
def analysis():
    return render_template("analysis.html")


@app.route("/manuals")
def manuals():
    return render_template("manuals.html")


@app.route("/master")
def master():
    return render_template("master_menu.html")

@app.route("/master/offices")
def office_master():
    return render_template(
        "office_master.html",
        offices=offices_for_current_company()
    )

@app.route("/master/offices/new", methods=["GET", "POST"])
def new_office():
    if request.method == "POST":
        name = request.form.get("name", "").strip()

        if Office.query.filter_by(
            company_code=session.get("company_code"),
            name=name
        ).first():
            return "この営業所はすでに登録されています。"

        office = Office(
            company_code=session.get("company_code"),
            name=name
        )

        db.session.add(office)
        db.session.commit()

        return redirect("/master/offices")

    return render_template(
        "office_form.html",
        office=None,
        index=None,
        mode="new"
    )


@app.route("/master/offices/<int:index>/edit", methods=["GET", "POST"])
def edit_office(index):
    office = Office.query.get(index)

    if not office:
        return redirect("/master/offices")

    if session.get("role") != "itc":
        if office.company_code != session.get("company_code"):
            return redirect("/master/offices")

    if request.method == "POST":
        office.name = request.form.get("name")
        db.session.commit()

        return redirect("/master/offices")

    office_dict = {
        "id": office.id,
        "index": office.id,
        "company_code": office.company_code,
        "name": office.name
    }

    return render_template(
        "office_form.html",
        office=office_dict,
        index=office.id,
        mode="edit"
    )


@app.route("/master/offices/<int:index>/delete", methods=["POST"])
def delete_office(index):
    office = Office.query.get(index)

    if not office:
        return redirect("/master/offices")

    if session.get("role") != "itc":
        if office.company_code != session.get("company_code"):
            return redirect("/master/offices")

    Driver.query.filter_by(
        company_code=office.company_code,
        office=office.name
    ).update(
        {"office": ""},
        synchronize_session=False
    )

    User.query.filter_by(
        company_code=office.company_code,
        office=office.name
    ).update(
        {"office": ""},
        synchronize_session=False
    )

    Vehicle.query.filter_by(
        company_code=office.company_code,
        office=office.name,
        deleted=False
    ).update(
        {"office": ""},
        synchronize_session=False
    )

    db.session.delete(office)
    db.session.commit()

    return redirect("/master/offices")

@app.route("/master/delivery-places")
def delivery_place_master():
    return render_template(
        "delivery_place_master.html",
        delivery_places=delivery_places_for_current_company()
    )


@app.route("/master/delivery-places/new", methods=["GET", "POST"])
def new_delivery_place():
    if request.method == "POST":
        name = request.form.get("name", "").strip()

        if DeliveryPlace.query.filter_by(
            company_code=session.get("company_code"),
            name=name
        ).first():
            return "この納入先はすでに登録されています。"

        place = DeliveryPlace(
            company_code=session.get("company_code"),
            name=name
        )

        db.session.add(place)
        db.session.commit()

        return redirect("/master/delivery-places")

    return render_template(
        "delivery_place_form.html",
        place=None,
        index=None,
        mode="new"
    )


@app.route("/master/delivery-places/<int:index>/edit", methods=["GET", "POST"])
def edit_delivery_place(index):
    place = DeliveryPlace.query.get(index)

    if not place:
        return redirect("/master/delivery-places")

    if session.get("role") != "itc":
        if place.company_code != session.get("company_code"):
            return redirect("/master/delivery-places")

    if request.method == "POST":
        place.name = request.form.get("name")
        db.session.commit()

        return redirect("/master/delivery-places")

    return render_template(
        "delivery_place_form.html",
        place={
            "id": place.id,
            "index": place.id,
            "company_code": place.company_code,
            "name": place.name
        },
        index=place.id,
        mode="edit"
    )


@app.route("/master/delivery-places/<int:index>/delete", methods=["POST"])
def delete_delivery_place(index):
    place = DeliveryPlace.query.get(index)

    if not place:
        return redirect("/master/delivery-places")

    if session.get("role") != "itc":
        if place.company_code != session.get("company_code"):
            return redirect("/master/delivery-places")

    db.session.delete(place)
    db.session.commit()

    return redirect("/master/delivery-places")

@app.route("/master/patrol-content-types")
def patrol_content_type_master():
    return render_template(
        "patrol_content_type_master.html",
        content_types=patrol_content_types_for_current_company()
    )


@app.route("/master/patrol-content-types/new", methods=["GET", "POST"])
def new_patrol_content_type():
    if request.method == "POST":
        name = request.form.get("name", "").strip()

        if PatrolContentType.query.filter_by(
            company_code=session.get("company_code"),
            name=name
        ).first():
            return "この内容区分はすでに登録されています。"

        item = PatrolContentType(
            company_code=session.get("company_code"),
            name=name
        )

        db.session.add(item)
        db.session.commit()

        return redirect("/master/patrol-content-types")

    return render_template(
        "patrol_content_type_form.html",
        item=None,
        mode="new"
    )


@app.route("/master/patrol-content-types/<int:index>/edit", methods=["GET", "POST"])
def edit_patrol_content_type(index):
    item = PatrolContentType.query.get(index)

    if not item:
        return redirect("/master/patrol-content-types")

    if session.get("role") != "itc":
        if item.company_code != session.get("company_code"):
            return redirect("/master/patrol-content-types")

    if request.method == "POST":
        item.name = request.form.get("name", "").strip()
        db.session.commit()

        return redirect("/master/patrol-content-types")

    return render_template(
        "patrol_content_type_form.html",
        item=item,
        mode="edit"
    )


@app.route("/master/patrol-content-types/<int:index>/delete", methods=["POST"])
def delete_patrol_content_type(index):
    item = PatrolContentType.query.get(index)

    if not item:
        return redirect("/master/patrol-content-types")

    if session.get("role") != "itc":
        if item.company_code != session.get("company_code"):
            return redirect("/master/patrol-content-types")

    db.session.delete(item)
    db.session.commit()

    return redirect("/master/patrol-content-types")

@app.route("/master/drivers")
def driver_master():
    keyword = request.args.get("keyword", "").strip()
    office = request.args.get("office", "").strip()
    vehicle = request.args.get("vehicle", "").strip()

    query = Driver.query.filter(
        Driver.company_code == session.get("company_code")
    )

    if keyword:
        keyword_like = f"%{keyword}%"

        query = query.filter(
            db.or_(
                Driver.employee_id.ilike(keyword_like),
                Driver.name.ilike(keyword_like)
            )
        )

    if office:
        query = query.filter(
            Driver.office == office
        )

    if vehicle:
        query = query.filter(
            Driver.vehicles_json.contains(f'"{vehicle}"')
        )

    driver_records = query.order_by(
        Driver.id.asc()
    ).all()

    filtered_drivers = []

    for driver in driver_records:
        driver_item = {
            "index": driver.id,
            "id": driver.id,
            "company_code": driver.company_code,
            "employee_id": driver.employee_id,
            "username": driver.employee_id,
            "name": driver.name,
            "role": driver.role,
            "office": driver.office,
            "safe_start_date": driver.safe_start_date,
            "vehicles": json.loads(driver.vehicles_json or "[]"),
            "licenses": json.loads(driver.licenses_json or "[]"),
        }

        safe_start_date = driver.safe_start_date

        if safe_start_date:
            start_date = datetime.strptime(
                safe_start_date,
                "%Y-%m-%d"
            )

            days = (datetime.today() - start_date).days
        else:
            days = 0

        years = days // 365
        remaining_days = days % 365

        driver_item["safe_days_display"] = (
            f"{years}年{remaining_days}日継続中"
            if years > 0
            else f"{days}日継続中"
        )

        filtered_drivers.append(driver_item)

    return render_template(
        "driver_master.html",
        drivers=filtered_drivers,
        offices=offices_for_current_company(),
        keyword=keyword,
        office=office,
        vehicle=vehicle,
    )
    
@app.route("/master/drivers/new", methods=["GET", "POST"])
def new_driver():
    if request.method == "POST":
        employee_id = request.form.get("employee_id")

        if User.query.filter_by(
            company_code=session.get("company_code"),
            username=employee_id
        ).first():
            return "このログインIDはすでに使用されています。"
        licenses = []

        license_types = request.form.getlist("license_type")
        license_expiries = request.form.getlist("license_expiry")

        for license_type, expiry in zip(license_types, license_expiries):
            if license_type and expiry:
                licenses.append({
                    "type": license_type,
                    "expiry": expiry
                })

        driver = Driver(
            company_code=session.get("company_code"),
            employee_id=request.form.get("employee_id"),
            name=request.form.get("name"),
            role=request.form.get("role"),
            office=request.form.get("office"),
            safe_start_date=request.form.get("safe_start_date"),
            vehicles_json=json.dumps(
                request.form.getlist("vehicles"),
                ensure_ascii=False
            ),
            licenses_json=json.dumps(
                licenses,
                ensure_ascii=False
            )
        )
        user = User(
            company_code=session.get("company_code"),
            username=request.form.get("employee_id"),
            password=generate_password_hash(request.form.get("password")),
            role=request.form.get("role"),
            name=request.form.get("name"),
            office=request.form.get("office"),
            favorite_vehicles_json=json.dumps(
                request.form.getlist("vehicles"),
                ensure_ascii=False
            )
        )

        db.session.add(user)

        db.session.add(driver)
        db.session.commit()

        return redirect("/master/drivers")

    return render_template(
        "driver_form.html",
        driver=None,
        offices=offices_for_current_company(),
        selected_vehicles=[],
        license_types=license_types_for_current_company(),
        mode="new"
    )

@app.route("/master/drivers/<int:index>/edit", methods=["GET", "POST"])
def edit_driver(index):
    driver = Driver.query.get(index)

    if not driver:
        return redirect("/master/drivers")

    if session.get("role") != "itc":
        if driver.company_code != session.get("company_code"):
            return redirect("/master/drivers")

    if request.method == "POST":
        licenses = []

        license_types = request.form.getlist("license_type")
        license_expiries = request.form.getlist("license_expiry")

        for license_type, expiry in zip(license_types, license_expiries):
            if license_type:
                licenses.append({
                    "type": license_type,
                    "expiry": expiry
                })

        old_employee_id = driver.employee_id

        new_employee_id = request.form.get("employee_id")

        if new_employee_id != old_employee_id:
            existing_user = User.query.filter_by(
                company_code=driver.company_code,
                username=new_employee_id
            ).first()

            if existing_user:
                return "このログインIDはすでに使用されています。"

        driver.employee_id = new_employee_id
        driver.name = request.form.get("name")
        driver.role = request.form.get("role")
        driver.office = request.form.get("office")
        driver.safe_start_date = request.form.get("safe_start_date")
        driver.vehicles_json = json.dumps(
            request.form.getlist("vehicles"),
            ensure_ascii=False
        )
        driver.licenses_json = json.dumps(
            licenses,
            ensure_ascii=False
        )
        user = User.query.filter_by(
            company_code=driver.company_code,
            username=old_employee_id
        ).first()

        if not user:
            user = User(
                company_code=driver.company_code,
                username=request.form.get("employee_id"),
                password=generate_password_hash(request.form.get("password") or "password"),
                role=request.form.get("role"),
                name=request.form.get("name"),
                office=request.form.get("office"),
                favorite_vehicles_json="[]"
            )
            db.session.add(user)

        user.username = request.form.get("employee_id")
        user.role = request.form.get("role")
        user.name = request.form.get("name")
        user.office = request.form.get("office")
        user.favorite_vehicles_json = json.dumps(
            request.form.getlist("vehicles"),
            ensure_ascii=False
        )

        if request.form.get("password"):
            user.password = generate_password_hash(request.form.get("password"))

        db.session.commit()

        return redirect("/master/drivers")

    driver_dict = driver_to_dict(driver)


    selected_vehicle_records = []

    if driver_dict["vehicles"]:

        selected_vehicle_records = Vehicle.query.filter(
            Vehicle.company_code == driver.company_code,
            Vehicle.vehicle_id.in_(driver_dict["vehicles"]),
            Vehicle.deleted == False
        ).all()


    selected_vehicles = []

    for vehicle in selected_vehicle_records:

        number = " ".join(
            value
            for value in [
                vehicle.plate_area or "",
                vehicle.plate_class or "",
                vehicle.plate_kana or "",
                vehicle.plate_number or "",
            ]
            if value
        )

        selected_vehicles.append({
            "vehicle_id": vehicle.vehicle_id,
            "number": number,
            "type": vehicle.type or "",
        })


    return render_template(
        "driver_form.html",
        driver=driver_dict,
        index=driver.id,
        offices=offices_for_current_company(),
        selected_vehicles=selected_vehicles,
        license_types=license_types_for_current_company(),
        mode="edit"
    )


@app.route("/master/drivers/<int:index>/delete", methods=["POST"])
def delete_driver(index):
    driver = Driver.query.get(index)

    if not driver:
        return redirect("/master/drivers")

    if session.get("role") != "itc":
        if driver.company_code != session.get("company_code"):
            return redirect("/master/drivers")

    user = User.query.filter_by(
        company_code=driver.company_code,
        username=driver.employee_id
    ).first()

    if user:
        db.session.delete(user)

    db.session.delete(driver)
    db.session.commit()

    return redirect("/master/drivers")

@app.route("/master/vehicles")
def vehicle_master():
    keyword = request.args.get("keyword", "").strip()
    office = request.args.get("office", "").strip()
    vehicle_type = request.args.get("vehicle_type", "").strip()
    status = request.args.get("status", "").strip()
    manufacturer = request.args.get("manufacturer", "").strip()
    model_code = request.args.get("model_code", "").strip()

    query = Vehicle.query.filter_by(
        company_code=session.get("company_code")
    )


    # キーワード検索
    if keyword:

        keyword_like = f"%{keyword}%"

        query = query.filter(
            db.or_(
                Vehicle.vehicle_id.ilike(keyword_like),
                Vehicle.plate_area.ilike(keyword_like),
                Vehicle.plate_class.ilike(keyword_like),
                Vehicle.plate_kana.ilike(keyword_like),
                Vehicle.plate_number.ilike(keyword_like),
                Vehicle.chassis_number.ilike(keyword_like),
                Vehicle.model_code.ilike(keyword_like),
                Vehicle.manufacturer.ilike(keyword_like),
                Vehicle.body_type.ilike(keyword_like),
            )
        )


    # 有効・無効
    if status == "active":
        query = query.filter(
            Vehicle.deleted == False
        )

    elif status == "inactive":
        query = query.filter(
            Vehicle.deleted == True
        )


    # メーカー・車名
    if manufacturer:
        query = query.filter(
            Vehicle.manufacturer == manufacturer
        )


    # 型式
    if model_code:
        query = query.filter(
            Vehicle.model_code == model_code
        )


    # 営業所
    if office:
        query = query.filter(
            Vehicle.office == office
        )


    # 車種
    if vehicle_type:
        query = query.filter(
            Vehicle.type == vehicle_type
        )


    # メーカー候補
    manufacturers = [
        item[0]
        for item in db.session.query(Vehicle.manufacturer)
        .filter(
            Vehicle.company_code == session.get("company_code"),
            Vehicle.manufacturer.isnot(None),
            Vehicle.manufacturer != ""
        )
        .distinct()
        .order_by(Vehicle.manufacturer)
        .all()
    ]


    # 型式候補
    model_codes = [
        item[0]
        for item in db.session.query(Vehicle.model_code)
        .filter(
            Vehicle.company_code == session.get("company_code"),
            Vehicle.model_code.isnot(None),
            Vehicle.model_code != ""
        )
        .distinct()
        .order_by(Vehicle.model_code)
        .all()
    ]


    # ページ分割
    page = request.args.get("page", 1, type=int)
    per_page = 100

    total_count = query.count()

    total_pages = max(
        1,
        (total_count + per_page - 1) // per_page
    )

    if page < 1:
        page = 1

    if page > total_pages:
        page = total_pages


    vehicle_records = (
        query
        .order_by(Vehicle.id.asc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )


    filtered_vehicles = []

    for vehicle in vehicle_records:

        item = {
            "index": vehicle.id,
            "id": vehicle.id,
            "company_code": vehicle.company_code,
            "vehicle_id": vehicle.vehicle_id,
            "deleted": vehicle.deleted,

            "plate_area": vehicle.plate_area,
            "plate_class": vehicle.plate_class,
            "plate_kana": vehicle.plate_kana,
            "plate_number": vehicle.plate_number,

            "chassis_number": vehicle.chassis_number,
            "model_code": vehicle.model_code,
            "first_registration_date": vehicle.first_registration_date,
            "manufacturer": vehicle.manufacturer,
            "body_type": vehicle.body_type,

            "gross_vehicle_weight": vehicle.gross_vehicle_weight,
            "max_payload": vehicle.max_payload,

            "type": vehicle.type,
            "office": vehicle.office,
            "inspection_expiry": vehicle.inspection_expiry,
        }

        item["number"] = vehicle_number(item)

        filtered_vehicles.append(item)

    company = Company.query.filter_by(
        company_code=session.get("company_code")
    ).first()

    vehicle_limit = 0

    vehicle_count = Vehicle.query.filter_by(
        company_code=session.get("company_code"),
        deleted=False
    ).count()

    remaining_vehicles = 0

    if company:
        vehicle_limit = company.vehicle_limit
        remaining_vehicles = vehicle_limit - vehicle_count

    return render_template(
        "vehicle_master.html",
        vehicles=filtered_vehicles,
        offices=offices_for_current_company(),
        vehicle_types=vehicle_types_for_current_company(),
        keyword=keyword,
        office=office,
        vehicle_type=vehicle_type,
        status=status,
        manufacturer=manufacturer,
        model_code=model_code,

        manufacturers=manufacturers,
        model_codes=model_codes,

        vehicle_limit=vehicle_limit,
        vehicle_count=vehicle_count,
        remaining_vehicles=remaining_vehicles,
        
        page=page,
        total_pages=total_pages,
        total_count=total_count,
    )

@app.route("/master/vehicles/import", methods=["GET", "POST"])
def import_vehicles():

    if request.method == "POST":

        excel_file = request.files.get("excel_file")

        if not excel_file or not excel_file.filename:
            return "Excelファイルを選択してください。"

        if not excel_file.filename.lower().endswith(".xlsx"):
            return "xlsx形式のExcelファイルを選択してください。"

        workbook = load_workbook(
            excel_file,
            data_only=True
        )

        sheet = workbook.active

        header_row = None

        for row in sheet.iter_rows():
            values = [cell.value for cell in row]

            if (
                "ID" in values
                and "車両番号" in values
                and "車体番号" in values
                and "車番" in values
            ):
                header_row = row[0].row
                break

        if header_row is None:
            return "車両台帳の見出し行が見つかりませんでした。"

        headers = [
            cell.value
            for cell in sheet[header_row]
        ]

        header_map = {
            name: index
            for index, name in enumerate(headers)
            if name is not None
        }

        vehicles_data = []

        for data_row in range(header_row + 1, sheet.max_row + 1):

            row_values = [
                cell.value
                for cell in sheet[data_row]
            ]

            # 空行は読み飛ばす
            if not any(value is not None for value in row_values):
                continue

            # 車番も車体番号も無い行は車両データではないとして読み飛ばす
            plate_number = row_values[header_map["車番"]]
            chassis_number = row_values[header_map["車体番号"]]

            if not plate_number and not chassis_number:
                continue

            vehicle_data = {
                "excel_row": data_row,

                "plate_area": row_values[header_map["車番・地域名"]],
                "plate_class": row_values[header_map["車番・分類"]],
                "plate_kana": row_values[header_map["車番・ひらがな"]],
                "plate_number": plate_number,

                "gross_vehicle_weight": row_values[header_map["車両総重量"]],
                "chassis_number": chassis_number,
                "model_code": row_values[header_map["車体型式"]],

                "first_registration_date": row_values[header_map["初年度車検"]],
                "inspection_expiry": row_values[header_map["車検期間終了日"]],

                "vehicle_name_raw": row_values[header_map["車両メーカー"]],
                "body_type_raw": row_values[header_map["車両形状"]],

                "vehicle_name": row_values[header_map["車両メーカー"]],
                "body_type": row_values[header_map["車両形状"]],

                "max_payload": row_values[header_map["最大積載量"]],
            }

            vehicles_data.append(vehicle_data)


        return render_template(
            "vehicle_import_preview.html",
            vehicles=vehicles_data,
            vehicle_types=vehicle_types_for_current_company(),
        )

    return render_template(
        "vehicle_import.html"
    )

@app.route("/master/vehicles/import/confirm", methods=["POST"])
def confirm_vehicle_import():

    company_code = session.get("company_code")

    plate_areas = request.form.getlist("plate_area")
    plate_classes = request.form.getlist("plate_class")
    plate_kanas = request.form.getlist("plate_kana")
    plate_numbers = request.form.getlist("plate_number")

    gross_weights = request.form.getlist("gross_vehicle_weight")
    chassis_numbers = request.form.getlist("chassis_number")
    model_codes = request.form.getlist("model_code")

    first_registration_dates = request.form.getlist(
        "first_registration_date"
    )

    inspection_expiries = request.form.getlist(
        "inspection_expiry"
    )

    vehicle_names = request.form.getlist("vehicle_name")
    body_types = request.form.getlist("body_type")
    max_payloads = request.form.getlist("max_payload")


    import_count = len(plate_numbers)

    company = Company.query.filter_by(
        company_code=company_code
    ).first()

    current_count = Vehicle.query.filter_by(
        company_code=company_code,
        deleted=False
    ).count()


    # 一括登録開始前から存在している車台番号
    import_chassis_numbers = {
        value.strip()
        for value in chassis_numbers
        if value and value.strip()
    }

    existing_chassis_numbers = set()

    if import_chassis_numbers:
        existing_chassis_numbers = {
            chassis_number.strip()
            for (chassis_number,) in db.session.query(
                Vehicle.chassis_number
            ).filter(
                Vehicle.company_code == company_code,
                Vehicle.deleted == False,
                Vehicle.chassis_number.in_(import_chassis_numbers)
            ).all()
            if chassis_number
        }


    # 上限チェック用
    # Excel内で一度数えた車台番号を記録
    counted_chassis_numbers = set()

    new_vehicle_count = 0


    for i in range(import_count):

        chassis_number = chassis_numbers[i].strip()


        # すでに登録済みなら新規台数に含めない
        if chassis_number:

            if chassis_number in existing_chassis_numbers:
                continue


            # Excel内重複も1台として数える
            if chassis_number in counted_chassis_numbers:
                continue


            counted_chassis_numbers.add(chassis_number)


        # 車台番号なしの行は、
        # 実際の登録処理と同じく1行＝1台として数える
        new_vehicle_count += 1


    # 新規登録予定台数で上限チェック
    if company:
        if current_count + new_vehicle_count > company.vehicle_limit:
            return (
                f"登録上限を超えます。"
                f"現在 {current_count} 台、"
                f"新規登録予定 {new_vehicle_count} 台、"
                f"上限 {company.vehicle_limit} 台です。"
            )


    max_number = db.session.query(
        db.func.max(
            db.cast(
                db.func.substr(Vehicle.vehicle_id, 2),
                db.Integer
            )
        )
    ).filter(
        Vehicle.company_code == company_code,
        Vehicle.vehicle_id.like("V%")
    ).scalar() or 0


    # 今回のExcel内ですでに処理した車台番号
    processed_chassis_numbers = set()

    registered_count = 0
    existing_skip_count = 0
    duplicate_skip_count = 0


    for i in range(import_count):

        chassis_number = chassis_numbers[i].strip()
        plate_area = plate_areas[i].strip()
        plate_class = plate_classes[i].strip()
        plate_kana = plate_kanas[i].strip()
        plate_number = plate_numbers[i].strip()


        # 一括登録開始前から存在していた車両
        if chassis_number and chassis_number in existing_chassis_numbers:
            existing_skip_count += 1
            continue


        # Excelファイル内で同じ車台番号が重複している場合
        if chassis_number and chassis_number in processed_chassis_numbers:
            duplicate_skip_count += 1
            continue


        if chassis_number:
            processed_chassis_numbers.add(chassis_number)


        max_number += 1

        new_id = f"V{max_number:03}"


        def to_int(value):
            value = str(value or "").replace(",", "").strip()

            if not value:
                return None

            return int(float(value))


        vehicle = Vehicle(
            company_code=company_code,
            vehicle_id=new_id,

            plate_area=plate_area,
            plate_class=plate_class,
            plate_kana=plate_kana,
            plate_number=plate_number,

            gross_vehicle_weight=to_int(
                gross_weights[i]
            ),

            chassis_number=chassis_number,
            model_code=model_codes[i].strip(),

            first_registration_date=
                first_registration_dates[i].strip(),

            inspection_expiry=
                inspection_expiries[i].strip(),

            # 画面上は「車名」だが、
            # 現在のDBでは manufacturer に保存
            manufacturer=vehicle_names[i].strip(),

            body_type=body_types[i].strip(),

            max_payload=to_int(
                max_payloads[i]
            ),

            # 今回のExcel登録対象外
            type="",
            office="",

            deleted=False,
        )

        db.session.add(vehicle)

        registered_count += 1


    db.session.commit()

    return render_template(
        "vehicle_import_complete.html",
        registered_count=registered_count,
        existing_skip_count=existing_skip_count,
        duplicate_skip_count=duplicate_skip_count,
    )
        
@app.route("/master/vehicles/new", methods=["GET", "POST"])
def new_vehicle():
    if request.method == "POST":
        company_code = session.get("company_code")



        vehicle_count = Vehicle.query.filter_by(
            company_code=company_code,
            deleted=False
        ).count()

        company = Company.query.filter_by(
            company_code=company_code
        ).first()

        if company:
            if vehicle_count >= company.vehicle_limit:
                return "登録可能台数の上限に達しています。"

        max_number = db.session.query(
            db.func.max(
                db.cast(
                    db.func.substr(Vehicle.vehicle_id, 2),
                    db.Integer
                )
            )
        ).filter(
            Vehicle.company_code == company_code,
            Vehicle.vehicle_id.like("V%")
        ).scalar() or 0

        new_id = f"V{max_number + 1:03}"

        vehicle = Vehicle(
            company_code=company_code,
            vehicle_id=new_id,

            plate_area=request.form.get("plate_area"),
            plate_class=request.form.get("plate_class"),
            plate_kana=request.form.get("plate_kana"),
            plate_number=request.form.get("plate_number"),

            chassis_number=request.form.get("chassis_number"),
            model_code=request.form.get("model_code"),
            first_registration_date=request.form.get("first_registration_date"),
            manufacturer=request.form.get("manufacturer"),
            body_type=request.form.get("body_type"),

            gross_vehicle_weight=int(request.form.get("gross_vehicle_weight") or 0),
            max_payload=int(request.form.get("max_payload") or 0),

            type=request.form.get("type"),
            office=request.form.get("office"),
            inspection_expiry=request.form.get("inspection_expiry"),
        )

        db.session.add(vehicle)
        db.session.commit()

        return redirect("/master/vehicles")

    return render_template(
        "vehicle_form.html",
        vehicle=None,
        offices=offices_for_current_company(),
        vehicle_types=vehicle_types_for_current_company(),
        mode="new"
    )


@app.route("/master/vehicles/<int:index>/edit", methods=["GET", "POST"])
def edit_vehicle(index):
    vehicle = Vehicle.query.get(index)

    if not vehicle:
        return redirect("/master/vehicles")

    if session.get("role") != "itc":
        if vehicle.company_code != session.get("company_code"):
            return redirect("/master/vehicles")

    if request.method == "POST":
        vehicle.plate_area = request.form.get("plate_area")
        vehicle.plate_class = request.form.get("plate_class")
        vehicle.plate_kana = request.form.get("plate_kana")
        vehicle.plate_number = request.form.get("plate_number")

        vehicle.chassis_number = request.form.get("chassis_number")
        vehicle.model_code = request.form.get("model_code")
        vehicle.first_registration_date = request.form.get("first_registration_date")
        vehicle.manufacturer = request.form.get("manufacturer")
        vehicle.body_type = request.form.get("body_type")

        vehicle.gross_vehicle_weight = int(
            request.form.get("gross_vehicle_weight") or 0
        )

        vehicle.max_payload = int(
            request.form.get("max_payload") or 0
        )

        vehicle.type = request.form.get("type")
        vehicle.office = request.form.get("office")
        vehicle.inspection_expiry = request.form.get("inspection_expiry")

        db.session.commit()

        return redirect("/master/vehicles")

    vehicle_dict = {
        "vehicle_id": vehicle.vehicle_id,
        "plate_area": vehicle.plate_area,
        "plate_class": vehicle.plate_class,
        "plate_kana": vehicle.plate_kana,
        "plate_number": vehicle.plate_number,
        "chassis_number": vehicle.chassis_number,
        "model_code": vehicle.model_code,
        "first_registration_date": vehicle.first_registration_date,
        "manufacturer": vehicle.manufacturer,
        "body_type": vehicle.body_type,
        "gross_vehicle_weight": vehicle.gross_vehicle_weight,
        "max_payload": vehicle.max_payload,
        "number": vehicle_number({
            "plate_area": vehicle.plate_area,
            "plate_class": vehicle.plate_class,
            "plate_kana": vehicle.plate_kana,
            "plate_number": vehicle.plate_number,
        }),
        "type": vehicle.type,
        "office": vehicle.office,
        "inspection_expiry": vehicle.inspection_expiry,
    }

    return render_template(
        "vehicle_form.html",
        vehicle=vehicle_dict,
        index=vehicle.id,
        offices=offices_for_current_company(),
        vehicle_types=vehicle_types_for_current_company(),
        mode="edit"
    )

@app.route("/master/vehicles/<int:index>/inactive", methods=["POST"])
def toggle_vehicle_inactive(index):

    vehicle = Vehicle.query.get(index)

    if not vehicle:
        return redirect("/master/vehicles")

    if session.get("role") != "itc":
        if vehicle.company_code != session.get("company_code"):
            return redirect("/master/vehicles")

    vehicle.deleted = request.form.get("inactive") == "1"

    db.session.commit()

    return redirect("/master/vehicles")

@app.route("/master/vehicles/bulk-delete", methods=["POST"])
def bulk_delete_vehicles():

    vehicle_indexes = request.form.getlist("vehicle_ids")

    if not vehicle_indexes:
        return redirect("/master/vehicles")

    company_code = session.get("company_code")

    vehicles_to_delete = Vehicle.query.filter(
        Vehicle.id.in_(vehicle_indexes),
        Vehicle.company_code == company_code
    ).all()

    if not vehicles_to_delete:
        return redirect("/master/vehicles")


    vehicle_ids_to_delete = {
        vehicle.vehicle_id
        for vehicle in vehicles_to_delete
        if vehicle.vehicle_id
    }
    
    delete_vehicle_patterns = [
        f'"{vehicle_id}"'
        for vehicle_id in vehicle_ids_to_delete
    ]


    # ドライバーの車両割当から削除
    driver_query = Driver.query.filter(
        Driver.company_code == company_code
    )

    if delete_vehicle_patterns:
        driver_query = driver_query.filter(
            db.or_(
                *[
                    Driver.vehicles_json.contains(pattern)
                    for pattern in delete_vehicle_patterns
                ]
            )
        )

    for driver in driver_query.all():

        driver_vehicles = json.loads(
            driver.vehicles_json or "[]"
        )

        new_driver_vehicles = [
            vehicle_id
            for vehicle_id in driver_vehicles
            if vehicle_id not in vehicle_ids_to_delete
        ]

        if new_driver_vehicles != driver_vehicles:

            driver.vehicles_json = json.dumps(
                new_driver_vehicles,
                ensure_ascii=False
            )


    # ユーザーのお気に入り車両から削除
    user_query = User.query.filter(
        User.company_code == company_code
    )

    if delete_vehicle_patterns:
        user_query = user_query.filter(
            db.or_(
                *[
                    User.favorite_vehicles_json.contains(pattern)
                    for pattern in delete_vehicle_patterns
                ]
            )
        )

    for user in user_query.all():

        favorite_vehicles = json.loads(
            user.favorite_vehicles_json or "[]"
        )

        new_favorite_vehicles = [
            vehicle_id
            for vehicle_id in favorite_vehicles
            if vehicle_id not in vehicle_ids_to_delete
        ]

        if new_favorite_vehicles != favorite_vehicles:

            user.favorite_vehicles_json = json.dumps(
                new_favorite_vehicles,
                ensure_ascii=False
            )


    # 関連履歴を削除してから車両本体を完全削除
    for vehicle in vehicles_to_delete:

        delete_vehicle_related_data(vehicle)

        db.session.delete(vehicle)


    db.session.commit()

    return redirect("/master/vehicles")

@app.route("/master/vehicles/bulk-inactive", methods=["POST"])
def bulk_inactive_vehicles():

    company_code = session.get("company_code")

    vehicle_indexes = request.form.getlist("vehicle_ids")

    if not vehicle_indexes:
        return redirect("/master/vehicles")

    Vehicle.query.filter(
        Vehicle.id.in_(vehicle_indexes),
        Vehicle.company_code == company_code
    ).update(
        {"deleted": True},
        synchronize_session=False
    )

    db.session.commit()

    return redirect("/master/vehicles")

@app.route("/master/vehicles/<int:index>/delete", methods=["POST"])
def delete_vehicle(index):
    vehicle = Vehicle.query.get(index)

    if not vehicle:
        return redirect("/master/vehicles")

    if session.get("role") != "itc":
        if vehicle.company_code != session.get("company_code"):
            return redirect("/master/vehicles")

    vehicle_id = vehicle.vehicle_id

    for driver in Driver.query.filter(
        Driver.company_code == vehicle.company_code,
        Driver.vehicles_json.contains(f'"{vehicle_id}"')
    ).all():
        vehicles = json.loads(driver.vehicles_json or "[]")

        if vehicle_id in vehicles:
            vehicles.remove(vehicle_id)
            driver.vehicles_json = json.dumps(
                vehicles,
                ensure_ascii=False
            )

    for user in User.query.filter(
        User.company_code == vehicle.company_code,
        User.favorite_vehicles_json.contains(f'"{vehicle_id}"')
    ).all():
        favorite_vehicles = json.loads(user.favorite_vehicles_json or "[]")

        if vehicle_id in favorite_vehicles:
            favorite_vehicles.remove(vehicle_id)
            user.favorite_vehicles_json = json.dumps(
                favorite_vehicles,
                ensure_ascii=False
            )

    # 関連履歴も完全削除
    delete_vehicle_related_data(vehicle)

    # 車両本体を完全削除
    db.session.delete(vehicle)

    db.session.commit()
    return redirect("/master/vehicles")

@app.route("/master/manuals")
def manual_master():
    return render_template(
        "manual_master.html",
        manuals=manuals_for_current_company()
    )

@app.route("/master/manuals/new", methods=["GET", "POST"])
def new_manual():
    if request.method == "POST":
        file = request.files.get("file")

        filename = save_uploaded_file(
            file,
            folder="static/manuals"
        )

        manual = Manual(
            company_code=session.get("company_code"),
            title=request.form.get("title"),
            category=request.form.get("category"),
            filename=filename
        )

        db.session.add(manual)
        db.session.commit()

        return redirect("/master/manuals")

    return render_template(
        "manual_form.html",
        manual=None,
        index=None,
        mode="new"
    )

@app.route("/master/manuals/<int:index>/edit", methods=["GET", "POST"])
def edit_manual(index):
    manual = Manual.query.get(index)

    if not manual:
        return redirect("/master/manuals")

    if session.get("role") != "itc":
        if manual.company_code != session.get("company_code"):
            return redirect("/master/manuals")

    if request.method == "POST":
        manual.title = request.form.get("title")
        manual.category = request.form.get("category")

        file = request.files.get("file")
        filename = save_uploaded_file(
            file,
            folder="static/manuals"
        )

        if filename:
            manual.filename = filename

        db.session.commit()

        return redirect("/master/manuals")

    manual_dict = {
        "id": manual.id,
        "index": manual.id,
        "company_code": manual.company_code,
        "title": manual.title,
        "category": manual.category,
        "filename": manual.filename,
    }

    return render_template(
        "manual_form.html",
        manual=manual_dict,
        index=manual.id,
        mode="edit"
    )

@app.route("/master/manuals/<int:index>/delete", methods=["POST"])
def delete_manual(index):
    manual = Manual.query.get(index)

    if not manual:
        return redirect("/master/manuals")

    if session.get("role") != "itc":
        if manual.company_code != session.get("company_code"):
            return redirect("/master/manuals")

    db.session.delete(manual)
    db.session.commit()

    return redirect("/master/manuals")

@app.route("/master/checklists")
def checklist_master():
    return render_template(
        "checklist_master.html",
        checklists=checklists_for_current_company()
    )

@app.route("/master/checklists/new", methods=["GET", "POST"])
def new_checklist():
    if request.method == "POST":
        item_categories = request.form.getlist("item_category")
        item_contents = request.form.getlist("item_content")
        input_types = request.form.getlist("input_type")
        item_types = request.form.getlist("item_type")
        approval_labels = request.form.getlist("approval_label")
        approval_allow_general_list = request.form.getlist("approval_allow_general")
        choices_list = request.form.getlist("choices")
        criteria_list = request.form.getlist("criteria")
        comment_required_list = request.form.getlist("comment_required")
        score_enabled = request.form.get("score_enabled") == "1"
        print_portrait = (
            request.form.get("target") == "車両管理"
            and request.form.get("print_portrait") == "1"
        )
        items = []

        for i in range(len(item_types)):
            item_type = item_types[i]
            
            if item_type == "inspector":
                items.append({
                    "item_type": "inspector",
                })
                continue

            if item_type == "approval":
                label = ""

                if i < len(approval_labels):
                    label = approval_labels[i]

                items.append({
                    "item_type": "approval",
                    "approval_label": label,
                    "approval_allow_general": str(i) in approval_allow_general_list,
                    "criteria_files": [],
                })

                continue

            if i >= len(item_contents):
                continue

            if not item_contents[i]:
                continue

            choices = []

            if input_types[i] == "select":
                choices = [
                    choice.strip()
                    for choice in choices_list[i].split(",")
                    if choice.strip()
                ]

            criteria_files = []

            for file in request.files.getlist(f"criteria_files_{i}"):
                filename = save_uploaded_file(file)

                if filename:
                    criteria_files.append(filename)

            items.append({
                "item_type": "check",
                "category": item_categories[i],
                "content": item_contents[i],
                "input_type": input_types[i],
                "choices": choices,
                "criteria": criteria_list[i],
                "criteria_files": criteria_files,
                "comment_required": str(i) in comment_required_list,
                "score_enabled": score_enabled,
            })

        frequency_value = ""
        frequency_unit = ""
        display_type = ""

        if request.form.get("target") == "車両管理":
            frequency_value = request.form.get("frequency_value")
            frequency_unit = request.form.get("frequency_unit")
            display_type = request.form.get("display_type")

        checklist = Checklist(
            company_code=session.get("company_code"),
            name=request.form.get("name"),
            target=request.form.get("target"),
            frequency_value=frequency_value,
            frequency_unit=frequency_unit,
            display_type=display_type,
            print_portrait=print_portrait,
            items_json=json.dumps(items, ensure_ascii=False)
        )

        db.session.add(checklist)
        db.session.commit()

        return redirect("/master/checklists")

    return render_template(
        "checklist_form.html",
        checklist=None,
        index=None,
        mode="new"
    )

@app.route("/safety/checklists")
def safety_checklists():
    safety_lists = []

    for checklist in checklists_for_current_company():
        if checklist["target"] == "安全管理":
            safety_lists.append(checklist)

    return render_template(
        "safety_checklists.html",
        checklists=safety_lists
    )

@app.route("/safety/checklists/<int:index>")
def safety_checklist_results(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/safety/checklists")

    if session.get("role") != "itc":
        if checklist_record.company_code != session.get("company_code"):
            return redirect("/safety/checklists")

    checklist = checklist_to_dict(checklist_record)

    results = []

    query = ChecklistResult.query.filter_by(
        company_code=session.get("company_code"),
        checklist_id=checklist_record.id
    )

    if session.get("role") != "itc":
        query = query.filter_by(
            company_code=session.get("company_code")
        )

    for result in query.order_by(ChecklistResult.id.desc()).all():
        item = checklist_result_to_dict(result)

        if not can_view_checklist_result(item):
            continue

        item["can_manage"] = can_manage_checklist_result(item)
        results.append(item)

    return render_template(
        "safety_checklist_results.html",
        checklist=checklist,
        checklist_index=checklist_record.id,
        results=results
    )


@app.route("/safety/checklist-results/<int:result_index>")
def checklist_result_detail(result_index):
    result_record = ChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/safety/checklists")

    result = checklist_result_to_dict(result_record)

    if not can_view_checklist_result(result):
        return redirect("/safety/checklists")

    checklist_record = Checklist.query.get(result_record.checklist_id)

    if not checklist_record:
        return redirect("/safety/checklists")

    checklist = checklist_to_dict(checklist_record)
    if not result.get("approvals"):
        approvals = []

        for item in checklist.get("items", []):
            if item.get("item_type") != "approval":
                continue

            approvals.append({
                "label": item.get("approval_label", ""),
                "allow_general": item.get("approval_allow_general", False),
                "approved_by": "",
                "approved_date": "",
            })

        result["approvals"] = approvals

    criteria_list = []

    for answer in result["answers"]:
        criteria = answer.get("criteria", "")
        if criteria and criteria not in criteria_list:
            criteria_list.append(criteria)

    summary = {}
    total_score = 0
    max_score = 0

    check_items = [
        item
        for item in checklist["items"]
        if item.get("item_type") != "approval"
    ]

    for item, answer in zip(check_items, result["answers"]):
        value = answer.get("value")

        # 自由記入は集計対象外
        if item.get("input_type") != "select":
            continue

        if value:
            if value not in summary:
                summary[value] = 0

            summary[value] += 1

        # 点数集計ONの場合だけ数値として計算
        if checklist.get("score_enabled"):
            try:
                total_score += float(value)
            except (TypeError, ValueError):
                pass

            numeric_choices = []

            for choice in item.get("choices", []):
                try:
                    numeric_choices.append(float(choice))
                except (TypeError, ValueError):
                    pass

            if numeric_choices:
                max_score += max(numeric_choices)

    return render_template(
        "checklist_result_detail.html",
        result=result,
        result_index=result_record.id,
        summary=summary,
        total_score=total_score,
        max_score=max_score,
        criteria_list=criteria_list,
        checklist=checklist,
        can_manage=can_manage_checklist_result(result),
        can_reject=can_reject_checklist_result(result),
    )

@app.route("/safety/checklist-results/<int:result_index>/excel")
def export_checklist_result_excel(result_index):

    result_record = ChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/safety/checklists")

    result = checklist_result_to_dict(result_record)

    if not can_view_checklist_result(result):
        return redirect("/safety/checklists")

    checklist_record = Checklist.query.get(result_record.checklist_id)

    if not checklist_record:
        return redirect("/safety/checklists")

    checklist = checklist_to_dict(checklist_record)
        # 評価基準
    criteria_list = []

    for answer in result["answers"]:
        criteria = answer.get("criteria", "")

        if criteria and criteria not in criteria_list:
            criteria_list.append(criteria)


    # チェック項目
    check_items = [
        item
        for item in checklist["items"]
        if item.get("item_type") != "approval"
    ]


    # 評価集計・合計点
    summary = {}
    total_score = 0
    max_score = 0

    for item, answer in zip(
        check_items,
        result["answers"]
    ):
        value = answer.get("value")

        # 自由記入は集計対象外
        if item.get("input_type") != "select":
            continue

        if value:
            summary[value] = summary.get(value, 0) + 1

        # 点数集計ONの場合だけ計算
        if checklist.get("score_enabled"):

            try:
                total_score += float(value)
            except (TypeError, ValueError):
                pass

            numeric_choices = []

            for choice in item.get("choices", []):
                try:
                    numeric_choices.append(
                        float(choice)
                    )
                except (TypeError, ValueError):
                    pass

            if numeric_choices:
                max_score += max(numeric_choices)


    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "チェックリスト結果"
    sheet.sheet_view.view = "pageBreakPreview"
    sheet.sheet_view.showGridLines = False

    # タイトル
    sheet.merge_cells("A1:L1")

    sheet["A1"] = f"{checklist['name']} 結果"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    sheet.row_dimensions[1].height = 28

    # 対象表示
    if result["target_type"] == "user":
        target_type_label = "ユーザー"
        target_value = result["target_user"] or "-"
    elif result["target_type"] == "vehicle":
        target_type_label = "車両"
        target_value = result["target_vehicle"] or "-"
    elif result["target_type"] == "office":
        target_type_label = "営業所"
        target_value = result["target_office"] or "-"
    else:
        target_type_label = "指定なし"
        target_value = ""

    info_rows = [
        ["実施日", result["checked_date"] or ""],
        ["点検者", result["checked_by"] or ""],
        ["対象", target_value or target_type_label],
        ["状態", result["status"] or ""],
    ]

    start_row = 3

    thin = Side(style="thin")

    for row_offset, values in enumerate(info_rows):
        row_number = start_row + row_offset

        # 見出し A:B
        sheet.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=2
        )

        label_cell = sheet.cell(
            row=row_number,
            column=1,
            value=values[0]
        )

        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        label_cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

        # 値 C:L
        sheet.merge_cells(
            start_row=row_number,
            start_column=3,
            end_row=row_number,
            end_column=12
        )

        value_cell = sheet.cell(
            row=row_number,
            column=3,
            value=values[1]
        )

        value_cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

        # 罫線
        for col_number in range(1, 13):
            sheet.cell(
                row=row_number,
                column=col_number
            ).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    # 上部情報の下から次の表示を開始
    current_row = start_row + len(info_rows) + 2


    # 評価基準
    if criteria_list:

        criteria_title_cell = sheet.cell(
            row=current_row,
            column=1,
            value="評価基準"
        )

        criteria_title_cell.font = Font(bold=True)

        criteria_title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        # 見出しはA:Hを結合
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=12
        )

        current_row += 1


        for criteria in criteria_list:

            sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=12
            )

            criteria_cell = sheet.cell(
                row=current_row,
                column=1,
                value=criteria
            )

            criteria_cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            for col_number in range(1, 13):
                sheet.cell(
                    row=current_row,
                    column=col_number
                ).border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

            current_row += 1


    # 評価集計
    if checklist.get("score_enabled") and summary:

        # 見出し A:B
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=2
        )

        summary_title_cell = sheet.cell(
            row=current_row,
            column=1,
            value="評価集計"
        )

        summary_title_cell.font = Font(bold=True)

        summary_title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        # 内容 C:H
        sheet.merge_cells(
            start_row=current_row,
            start_column=3,
            end_row=current_row,
            end_column=12
        )

        summary_text = " / ".join(
            f"{value}：{count}件"
            for value, count in summary.items()
        )

        sheet.cell(
            row=current_row,
            column=3,
            value=summary_text
        )

        for col_number in range(1, 13):
            sheet.cell(
                row=current_row,
                column=col_number
            ).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

        current_row += 1
        
    # 合計点
    if checklist.get("score_enabled"):

        # 見出し A:B
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=2
        )

        score_title_cell = sheet.cell(
            row=current_row,
            column=1,
            value="合計点"
        )

        score_title_cell.font = Font(bold=True)

        score_title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        # 点数 C:H
        sheet.merge_cells(
            start_row=current_row,
            start_column=3,
            end_row=current_row,
            end_column=12
        )

        sheet.cell(
            row=current_row,
            column=3,
            value=(
                f"{int(total_score)} / {int(max_score)}点"
                if max_score
                else f"{int(total_score)}点"
            )
        )

        for col_number in range(1, 13):
            sheet.cell(
                row=current_row,
                column=col_number
            ).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

        current_row += 2

    else:
        current_row += 1

    # チェック結果
    result_header_row = current_row
    sheet.row_dimensions[result_header_row].height = 24

    # A:B = カテゴリ
    # C:G = チェック内容
    # H   = 評価
    # I:L = コメント
    header_ranges = [
        ("カテゴリ", 1, 2),
        ("チェック内容", 3, 7),
        ("評価", 8, 8),
        ("コメント", 9, 12),
    ]

    for header, start_col, end_col in header_ranges:

        sheet.merge_cells(
            start_row=current_row,
            start_column=start_col,
            end_row=current_row,
            end_column=end_col
        )

        cell = sheet.cell(
            row=current_row,
            column=start_col,
            value=header
        )

        cell.font = Font(bold=True)

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        for col_number in range(
            start_col,
            end_col + 1
        ):
            sheet.cell(
                row=current_row,
                column=col_number
            ).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    current_row += 1


    for item, answer in zip(
        check_items,
        result["answers"]
    ):

        # カテゴリ A:B
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=2
        )

        sheet.cell(
            row=current_row,
            column=1,
            value=answer.get("category", "")
        )

        # チェック内容 C:G
        sheet.merge_cells(
            start_row=current_row,
            start_column=3,
            end_row=current_row,
            end_column=7
        )

        sheet.cell(
            row=current_row,
            column=3,
            value=answer.get("content", "")
        )

        # コメント I:L
        sheet.merge_cells(
            start_row=current_row,
            start_column=9,
            end_row=current_row,
            end_column=12
        )

        if item.get("input_type") == "select":

            sheet.cell(
                row=current_row,
                column=8,
                value=answer.get("value", "")
            )

            sheet.cell(
                row=current_row,
                column=9,
                value=answer.get("comment", "") or ""
            )

        else:

            sheet.cell(
                row=current_row,
                column=9,
                value=answer.get("value", "") or ""
            )

        # 罫線
        for col_number in range(1, 13):
            sheet.cell(
                row=current_row,
                column=col_number
            ).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

        current_row += 1
    
    # 12列構成
    # A～Lはすべて同じ幅
    for column_letter in [
        "A", "B", "C", "D", "E", "F",
        "G", "H", "I", "J", "K", "L"
    ]:
        sheet.column_dimensions[column_letter].width = 9

    # 折り返し
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical or "top",
                wrap_text=True
            )


    import math
    import unicodedata


    def text_width(text):
        width = 0

        for char in str(text or ""):
            if unicodedata.east_asian_width(char) in ("W", "F", "A"):
                width += 2
            else:
                width += 1

        return width


    def wrapped_line_count(text, available_width):
        text = str(text or "")

        if not text:
            return 1

        line_count = 0

        for line in text.split("\n"):
            line_count += max(
                1,
                math.ceil(
                    text_width(line) / available_width
                )
            )

        return line_count


    # チェック項目の文字量に応じて行高を調整
    for row_number in range(
        result_header_row + 1,
        current_row
    ):

        # チェック内容 C:G
        content = sheet.cell(
            row=row_number,
            column=3
        ).value or ""

        # コメント I:L
        comment = sheet.cell(
            row=row_number,
            column=9
        ).value or ""

        content_lines = wrapped_line_count(
            content,
            52
        )

        comment_lines = wrapped_line_count(
            comment,
            41
        )

        line_count = max(
            content_lines,
            comment_lines
        )

        sheet.row_dimensions[row_number].height = max(
            28,
            (line_count * 22) + 6
        )

    # 評価列 H を中央揃え
    for row_number in range(
        result_header_row + 1,
        current_row
    ):
        sheet.cell(
            row=row_number,
            column=8
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
    
    # 承認・押印欄
    approval_items = [
        item
        for item in checklist["items"]
        if item.get("item_type") == "approval"
    ]

    approval_results = result.get("approvals", [])

    stamp_entries = []

    for approval_index, item in enumerate(approval_items):
        label = item.get("approval_label", "").strip()

        if not label:
            continue

        approval_result = (
            approval_results[approval_index]
            if approval_index < len(approval_results)
            else {}
        )

        stamp_entries.append({
            "label": label,
            "approved_by": approval_result.get("approved_by", ""),
            "approved_date": approval_result.get("approved_date", ""),
        })

    stamp_headers = [
        entry["label"]
        for entry in stamp_entries
    ]

    stamp_start_row = current_row + 2

    if stamp_headers:

        # 承認欄タイトル
        stamp_title_cell = sheet.cell(
            row=stamp_start_row,
            column=12,
            value="確認・押印"
        )

        stamp_title_cell.font = Font(bold=True)

        stamp_title_cell.alignment = Alignment(
            horizontal="right",
            vertical="center"
        )

        # 1行につき最大6つ
        entry_chunks = [
            stamp_entries[i:i + 6]
            for i in range(0, len(stamp_entries), 6)
        ]

        row_number = stamp_start_row + 1

        for entry_chunk in entry_chunks:

            item_count = len(entry_chunk)

            stamp_header_row = row_number
            stamp_box_row = row_number + 1

            # 押印欄はセル結合しない
            # A～Lは全列同じ幅
            # 1項目につき1セル、間に1セル空けて均等配置
            #
            # 1個 → L
            # 2個 → J / L
            # 3個 → H / J / L
            # 4個 → F / H / J / L
            # 5個 → D / F / H / J / L
            # 6個 → B / D / F / H / J / L

            all_stamp_columns = [7, 8, 9, 10, 11, 12]

            stamp_columns = all_stamp_columns[
                6 - item_count:
            ]

            for entry, column in zip(
                entry_chunk,
                stamp_columns
            ):

                # 見出し
                header_cell = sheet.cell(
                    row=stamp_header_row,
                    column=column,
                    value=entry["label"]
                )

                header_cell.font = Font(bold=True)

                header_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                header_cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

                # 押印内容
                stamp_value = entry["approved_by"]

                if entry["approved_by"] and entry["approved_date"]:
                    stamp_value += (
                        f"\n{entry['approved_date']}"
                    )

                stamp_cell = sheet.cell(
                    row=stamp_box_row,
                    column=column,
                    value=stamp_value
                )

                stamp_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                stamp_cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

            sheet.row_dimensions[
                stamp_header_row
            ].height = 22

            sheet.row_dimensions[
                stamp_box_row
            ].height = 55

            row_number = stamp_box_row + 1

        current_row = row_number
                
    # A4印刷設定
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    
    # 印刷時に水平方向の中央へ配置
    sheet.print_options.horizontalCentered = True

    # 印刷範囲
    sheet.print_area = f"A1:L{current_row - 1}"

    # 余白
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2

    # 2ページ目以降もチェック結果の見出しを表示
    sheet.print_title_rows = (
        f"{result_header_row}:{result_header_row}"
    )
    
            
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    safe_checklist_name = checklist["name"].replace("/", "_").replace("\\", "_")
    safe_checked_date = (
        (result["checked_date"] or "")
        .replace("/", "-")
        .replace(":", "-")
    )

    filename = (
        f"{safe_checklist_name}_"
        f"{safe_checked_date}_"
        f"結果.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    
@app.route("/safety/checklist-results/<int:result_index>/edit", methods=["GET", "POST"])
def edit_checklist_result(result_index):
    result_record = ChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/safety/checklists")

    result = checklist_result_to_dict(result_record)

    if not can_manage_checklist_result(result):
        return redirect(f"/safety/checklist-results/{result_index}")

    checklist_record = Checklist.query.get(result_record.checklist_id)

    if not checklist_record:
        return redirect("/safety/checklists")

    checklist = checklist_to_dict(checklist_record)

    if request.method == "POST":
        answers = []
        answer_index = 0

        for item in checklist["items"]:
            if item.get("item_type") == "approval":
                continue

            answers.append({
                "category": item.get("category", ""),
                "content": item.get("content", ""),
                "criteria": item.get("criteria", ""),
                "criteria_files": item.get("criteria_files", []),
                "value": request.form.get(f"answer_{answer_index}"),
                "comment": request.form.get(f"comment_{answer_index}"),
                "files": result["answers"][answer_index].get("files", []) if answer_index < len(result["answers"]) else [],
                "patrol_link": False,
            })

            answer_index += 1

        result_record.target_type = request.form.get("target_type")
        result_record.target_user = request.form.get("target_user")
        result_record.target_vehicle = request.form.get("target_vehicle")
        result_record.target_office = request.form.get("target_office")
        result_record.answers_json = json.dumps(answers, ensure_ascii=False)

        db.session.commit()

        return redirect(f"/safety/checklist-results/{result_record.id}")

    criteria_list = []

    for item in checklist["items"]:
        criteria = item.get("criteria", "")

        if criteria and criteria not in criteria_list:
            criteria_list.append(criteria)

    selected_vehicle = None

    if result.get("target_vehicle"):

        vehicle_record = Vehicle.query.filter_by(
            company_code=session.get("company_code"),
            vehicle_id=result.get("target_vehicle")
        ).first()

        if vehicle_record:

            number = " ".join(
                value
                for value in [
                    vehicle_record.plate_area or "",
                    vehicle_record.plate_class or "",
                    vehicle_record.plate_kana or "",
                    vehicle_record.plate_number or "",
                ]
                if value
            )

            selected_vehicle = {
                "vehicle_id": vehicle_record.vehicle_id,
                "number": number,
                "manufacturer": vehicle_record.manufacturer or "",
                "model_code": vehicle_record.model_code or "",
            }
    return render_template(
        "checklist_result_form.html",
        checklist=checklist,
        index=checklist_record.id,
        result=result,
        result_index=result_record.id,
        mode="edit",
        criteria_list=criteria_list,
        drivers=drivers_for_current_company(),
        selected_vehicle=selected_vehicle,
        offices=offices_for_current_company()
    )

@app.route("/safety/checklist-results/<int:result_index>/delete", methods=["POST"])
def delete_checklist_result(result_index):
    result_record = ChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/safety/checklists")

    result = checklist_result_to_dict(result_record)

    if not can_manage_checklist_result(result):
        return redirect(f"/safety/checklist-results/{result_index}")

    checklist_id = result_record.checklist_id

    db.session.delete(result_record)
    db.session.commit()

    return redirect(f"/safety/checklists/{checklist_id}")

@app.route("/vehicle/checklists")
def vehicle_checklists():
    vehicle_lists = []

    for checklist in checklists_for_current_company():
        if checklist["target"] == "車両管理":
            vehicle_lists.append(checklist)

    return render_template(
        "vehicle_checklists.html",
        checklists=vehicle_lists
    )

@app.route("/vehicle/checklists/<int:index>")
def vehicle_checklist_results(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/vehicle/checklists")

    if checklist_record.company_code != session.get("company_code"):
        return redirect("/vehicle/checklists")

    checklist = checklist_to_dict(checklist_record)

    if checklist["target"] != "車両管理":
        return redirect("/vehicle/checklists")

    year = request.args.get("year", str(datetime.now().year))
    month = request.args.get("month", str(datetime.now().month).zfill(2))
    vehicle_id = request.args.get("vehicle_id", "")

    if checklist.get("frequency_unit") == "year":
        default_active_day = str(datetime.now().year)
    elif checklist.get("display_type") == "month":
        default_active_day = str(datetime.now().day).zfill(2)
    else:
        default_active_day = str(datetime.now().month).zfill(2)

    active_day = request.args.get("active_day", default_active_day)

    if not active_day:
        active_day = datetime.now().strftime("%d")

    display_days = []
    display_weekdays = {}
    week_names = ["月", "火", "水", "木", "金", "土", "日"]

    frequency_unit = checklist.get("frequency_unit", "")
    display_mode = "month"

    if frequency_unit == "year":
        display_mode = "year_list"
        base_year = int(year)

        for y in range(base_year, base_year + 5):
            display_days.append(y)
            display_weekdays[y] = {
                "name": "",
                "is_weekend": False
            }

    elif checklist.get("display_type") == "month":
        display_mode = "day_list"

        import calendar
        last_day = calendar.monthrange(int(year), int(month))[1]

        for day in range(1, last_day + 1):
            display_days.append(day)

            weekday_index = datetime(int(year), int(month), day).weekday()

            display_weekdays[day] = {
                "name": week_names[weekday_index],
                "is_weekend": weekday_index in [5, 6]
            }

    else:
        display_mode = "month_list"

        for m in range(1, 13):
            display_days.append(m)
            display_weekdays[m] = {
                "name": "",
                "is_weekend": False
            }

    results = []

    query = VehicleChecklistResult.query.filter_by(
        company_code=checklist_record.company_code,
        checklist_id=checklist_record.id
    )

    if vehicle_id:
        query = query.filter_by(
            vehicle_id=vehicle_id
        )

    for result_record in query.all():
        result = vehicle_checklist_result_to_dict(result_record)

        if display_mode == "year_list":
            if int(result.get("year", 0)) not in [int(d) for d in display_days]:
                continue

        elif display_mode == "month_list":
            if str(result.get("year")) != str(year):
                continue

        else:
            if str(result.get("year")) != str(year):
                continue

            if str(result.get("month")).zfill(2) != str(month).zfill(2):
                continue

        results.append(result)
    selected_notify_users = []

    for result in results:
        if display_mode == "year_list":
            is_active_result = (
                str(result.get("year")) == str(active_day)
            )

        elif display_mode == "month_list":
            is_active_result = (
                str(result.get("month")).zfill(2)
                == str(active_day).zfill(2)
            )

        else:
            is_active_result = (
                str(result.get("day")).zfill(2)
                == str(active_day).zfill(2)
            )

        if is_active_result:
            selected_notify_users = result.get("notify_users", [])
            break

    selected_vehicle = None

    if vehicle_id:

        vehicle_record = Vehicle.query.filter_by(
            company_code=checklist_record.company_code,
            vehicle_id=vehicle_id
        ).first()

        if vehicle_record:

            number = " ".join(
                value
                for value in [
                    vehicle_record.plate_area or "",
                    vehicle_record.plate_class or "",
                    vehicle_record.plate_kana or "",
                    vehicle_record.plate_number or "",
                ]
                if value
            )

            selected_vehicle = {
                "vehicle_id": vehicle_record.vehicle_id,
                "number": number,
                "manufacturer": vehicle_record.manufacturer or "",
                "model_code": vehicle_record.model_code or "",
            }
                    
    return render_template(
        "vehicle_checklist_results.html",
        checklist=checklist,
        checklist_index=checklist_record.id,
        results=results,
        selected_notify_users=selected_notify_users,
        selected_vehicle=selected_vehicle,
        year=year,
        month=month,
        vehicle_id=vehicle_id,
        active_day=active_day,
        display_days=display_days,
        display_weekdays=display_weekdays,
        display_mode=display_mode
    )


@app.route(
    "/vehicle/checklist-results/<int:result_index>/approve/<int:approval_index>",
    methods=["POST"]
)
def approve_vehicle_checklist_result(result_index, approval_index):
    result_record = VehicleChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/vehicle/checklists")

    if result_record.company_code != session.get("company_code"):
        return redirect("/vehicle/checklists")

    approvals = json.loads(
        result_record.approvals_json or "[]"
    )
    if not approvals:
        checklist_record = Checklist.query.get(
            result_record.checklist_id
        )

        if checklist_record:
            checklist = checklist_to_dict(checklist_record)

            for item in checklist.get("items", []):
                if item.get("item_type") != "approval":
                    continue

                approvals.append({
                    "label": item.get("approval_label", ""),
                    "allow_general": item.get("approval_allow_general", False),
                    "approved_by": "",
                    "approved_date": "",
                })

    if approval_index < 0 or approval_index >= len(approvals):
        return redirect("/vehicle/checklists")

    approval = approvals[approval_index]
    
    # 旧データに allow_general が無い場合は
    # 現在のチェックリストマスタから補完
    if "allow_general" not in approval:
        checklist_record = Checklist.query.get(
            result_record.checklist_id
        )

        if checklist_record:
            checklist = checklist_to_dict(checklist_record)

            approval_items = [
                item
                for item in checklist.get("items", [])
                if item.get("item_type") == "approval"
            ]

            if approval_index < len(approval_items):
                approval["allow_general"] = approval_items[
                    approval_index
                ].get(
                    "approval_allow_general",
                    False
                )
                
    result = vehicle_checklist_result_to_dict(result_record)

    if not can_approve_checklist_result(result, approval):
        return redirect("/vehicle/checklists")
    
    approval["approved_by"] = session.get("name")
    approval["approved_date"] = datetime.now().strftime(
        "%Y-%m-%d %H:%M"
    )

    result_record.approvals_json = json.dumps(
        approvals,
        ensure_ascii=False
    )

    result_record.reject_reason = ""

    all_approved = all(
        item.get("approved_by")
        for item in approvals
    )

    if approvals and all_approved:
        result_record.status = "承認済み"
        result_record.approved_by = session.get("name")
        result_record.approved_date = datetime.now().strftime(
            "%Y-%m-%d %H:%M"
        )
    else:
        result_record.status = "承認待ち"
        result_record.approved_by = ""
        result_record.approved_date = ""

    db.session.commit()

    checklist_record = Checklist.query.get(
        result_record.checklist_id
    )

    if checklist_record:
        checklist = checklist_to_dict(checklist_record)

        if checklist.get("frequency_unit") == "year":
            active_value = result_record.year

        elif checklist.get("display_type") == "month":
            active_value = result_record.day

        else:
            active_value = result_record.month

    else:
        active_value = result_record.day

    return redirect(
        f"/vehicle/checklists/{result_record.checklist_id}"
        f"?vehicle_id={result_record.vehicle_id}"
        f"&year={result_record.year}"
        f"&month={result_record.month}"
        f"&active_day={active_value}"
    )

@app.route("/vehicle/checklist-results/<int:result_index>/excel")
def export_vehicle_checklist_result_excel(result_index):
    result_record = VehicleChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/vehicle/checklists")

    if session.get("role") != "itc":
        if result_record.company_code != session.get("company_code"):
            return redirect("/vehicle/checklists")

    result = vehicle_checklist_result_to_dict(result_record)

    checklist_record = Checklist.query.get(
        result_record.checklist_id
    )

    if not checklist_record:
        return redirect("/vehicle/checklists")

    checklist = checklist_to_dict(checklist_record)
    # 点検頻度
    try:
        frequency_value = int(
            checklist.get("frequency_value") or 1
        )
    except (TypeError, ValueError):
        frequency_value = 1

    frequency_unit = checklist.get(
        "frequency_unit",
        ""
    )

    display_type = checklist.get(
        "display_type",
        ""
    )
    
    # Excelの表示単位を決定
    if frequency_unit == "year":
        excel_display_mode = "year"

    elif display_type == "month":
        excel_display_mode = "day"

    else:
        excel_display_mode = "month"
        
    vehicle_record = Vehicle.query.filter_by(
        company_code=result_record.company_code,
        vehicle_id=result_record.vehicle_id
    ).first()

    vehicle_info = {
        "vehicle_id": result_record.vehicle_id or "",
        "number": "",
        "manufacturer": "",
        "model_code": "",
    }

    if vehicle_record:
        vehicle_info["number"] = " ".join(
            value
            for value in [
                vehicle_record.plate_area or "",
                vehicle_record.plate_class or "",
                vehicle_record.plate_kana or "",
                vehicle_record.plate_number or "",
            ]
            if value
        )

        vehicle_info["manufacturer"] = (
            vehicle_record.manufacturer or ""
        )

        vehicle_info["model_code"] = (
            vehicle_record.model_code or ""
        )
    
    # 点検頻度に応じて出力対象を取得
    result_query = VehicleChecklistResult.query.filter_by(
        company_code=result_record.company_code,
        checklist_id=result_record.checklist_id,
        vehicle_id=result_record.vehicle_id
    )

    if frequency_unit == "year":
        # 年次・複数年ごとの点検
        result_records = result_query.order_by(
            VehicleChecklistResult.year.asc()
        ).all()

    elif display_type == "month":
        # 日次系
        result_records = result_query.filter_by(
            year=result_record.year,
            month=result_record.month
        ).order_by(
            VehicleChecklistResult.day.asc()
        ).all()

    else:
        # 月例・3か月ごと・6か月ごと等
        result_records = result_query.filter_by(
            year=result_record.year
        ).order_by(
            VehicleChecklistResult.month.asc()
        ).all()

    period_results = [
        vehicle_checklist_result_to_dict(record)
        for record in result_records
    ]
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "車両点検結果"

    # 印刷設定
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    if checklist.get("print_portrait"):
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    else:
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_options.horizontalCentered = True

    # 点検頻度に応じた列幅
    sheet.column_dimensions["A"].width = 42

    if excel_display_mode == "day":
        days_in_month = calendar.monthrange(
            int(result_record.year),
            int(result_record.month)
        )[1]

        end_column = days_in_month + 1
        period_width = 3.8

    elif excel_display_mode == "month":
        end_column = 13
        period_width = 6

    else:
        end_column = 6
        period_width = 12

    for column in range(2, end_column + 1):
        sheet.column_dimensions[
            sheet.cell(row=1, column=column).column_letter
        ].width = period_width

    # タイトル
    title_end_column = end_column

    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=2,
        end_column=title_end_column
    )

    title_cell = sheet["A1"]
    if excel_display_mode == "day":
        title_cell.value = (
            f"{result_record.year}年 "
            f"{int(result_record.month)}月 "
            f"{checklist.get('name', '車両点検表')}"
        )

    elif excel_display_mode == "month":
        title_cell.value = (
            f"{result_record.year}年 "
            f"{checklist.get('name', '車両点検表')}"
        )

    else:
        title_cell.value = (
            f"{checklist.get('name', '車両点検表')}"
        )
    title_cell.font = Font(
        bold=True,
        size=16
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # 評価基準
    criteria_list = []

    for item in checklist.get("items", []):
        if item.get("item_type") != "check":
            continue

        criteria = (item.get("criteria") or "").strip()

        if criteria and criteria not in criteria_list:
            criteria_list.append(criteria)

    if criteria_list:
        sheet.row_dimensions[3].height = 24
        
        sheet.merge_cells(
            start_row=3,
            start_column=1,
            end_row=3,
            end_column=end_column
        )

        criteria_cell = sheet.cell(
            row=3,
            column=1,
            value="評価基準：" + " / ".join(criteria_list)
        )

        criteria_cell.font = Font(
            size=9
        )

        criteria_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )
    
    # 月間帳票の車両情報
    vehicle_number = (
        vehicle_info["number"]
        or vehicle_info["vehicle_id"]
    )

    sheet.merge_cells(
        start_row=4,
        start_column=1,
        end_row=4,
        end_column=end_column
    )

    sheet["A4"] = f"車番　{vehicle_number}"
    sheet["A4"].font = Font(bold=True)
    sheet["A4"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )
    
    # 点検表ヘッダー
    header_row = 5
    weekday_row = 6

    sheet["A5"] = "点検項目"

    sheet.merge_cells(
        start_row=5,
        start_column=1,
        end_row=6,
        end_column=1
    )

    year = int(result_record.year)
    month = int(result_record.month)

    if excel_display_mode == "day":
        # 日次：1～31日
        days_in_month = calendar.monthrange(
            year,
            month
        )[1]

        weekday_names = [
            "月", "火", "水", "木", "金", "土", "日"
        ]

        for day in range(1, days_in_month + 1):
            column = day + 1

            if day <= days_in_month:
                sheet.cell(
                    row=header_row,
                    column=column,
                    value=day
                )

                weekday_index = datetime(
                    year,
                    month,
                    day
                ).weekday()

                sheet.cell(
                    row=weekday_row,
                    column=column,
                    value=weekday_names[weekday_index]
                )

                if weekday_index == 5:
                    sheet.cell(
                        row=weekday_row,
                        column=column
                    ).font = Font(color="0000FF")

                elif weekday_index == 6:
                    sheet.cell(
                        row=weekday_row,
                        column=column
                    ).font = Font(color="FF0000")
                    
    elif excel_display_mode == "month":
        # 月例・3か月・6か月ごと等：1～12月
        for month_no in range(1, 13):
            column = month_no + 1

            sheet.cell(
                row=header_row,
                column=column,
                value=f"{month_no}月"
            )

            sheet.merge_cells(
                start_row=header_row,
                start_column=column,
                end_row=weekday_row,
                end_column=column
            )

    else:
        # 年次・複数年ごと
        base_year = int(result_record.year)

        for offset in range(5):
            column = offset + 2
            target_year = base_year + offset

            sheet.cell(
                row=header_row,
                column=column,
                value=f"{target_year}年"
            )

            sheet.merge_cells(
                start_row=header_row,
                start_column=column,
                end_row=weekday_row,
                end_column=column
            )

    # 点検項目を縦に並べる
    current_row = 7
    previous_category = None
    category_rows = set()

    for item_no, item in enumerate(checklist.get("items", [])):
        
        if item.get("item_type") != "check":
            continue

        category = item.get("category", "").strip()
        content = item.get("content", "").strip()

        # カテゴリが変わったら横一列のカテゴリ行を追加
        if category and category != previous_category:
            category_row = current_row
            category_rows.add(category_row)

            sheet.merge_cells(
                start_row=category_row,
                start_column=1,
                end_row=category_row,
                end_column=end_column
            )

            category_cell = sheet.cell(
                row=category_row,
                column=1,
                value=f"（{category}）"
            )

            category_cell.font = Font(bold=True)
            category_cell.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

            sheet.row_dimensions[category_row].height = 20

            current_row += 1
            previous_category = category

        # 点検項目にはカテゴリ名を付けない
        item_text = content

        if "\n" not in item_text and len(item_text) <= 35:
            sheet.row_dimensions[current_row].height = 24
            
        sheet.cell(
            row=current_row,
            column=1,
            value=item_text
        )

        sheet.cell(
            row=current_row,
            column=1
        ).alignment = Alignment(
            vertical="center",
            wrap_text=True
        )
        
        # 点検結果を表示単位に応じて入れる
        for period_result in period_results:

            if excel_display_mode == "day":
                try:
                    period_value = int(
                        period_result.get("day", 0)
                    )
                except (TypeError, ValueError):
                    continue

                if period_value < 1 or period_value > 31:
                    continue

                column = period_value + 1

            elif excel_display_mode == "month":
                try:
                    period_value = int(
                        period_result.get("month", 0)
                    )
                except (TypeError, ValueError):
                    continue

                if period_value < 1 or period_value > 12:
                    continue

                column = period_value + 1

            else:
                try:
                    period_year = int(
                        period_result.get("year", 0)
                    )
                except (TypeError, ValueError):
                    continue

                year_offset = period_year - base_year

                if year_offset < 0 or year_offset >= 5:
                    continue

                column = year_offset + 2

            matched_answer = None

            for answer in period_result.get("answers", []):
                answer_item_no = answer.get("item_no")

                # item_no があるデータは番号だけで照合
                if answer_item_no not in (None, ""):
                    if str(answer_item_no) == str(item_no):
                        matched_answer = answer
                        break
                    continue

                # item_no が無い旧データだけカテゴリー＋内容で照合
                if (
                    answer.get("category", "")
                    == item.get("category", "")
                    and answer.get("content", "")
                    == item.get("content", "")
                ):
                    matched_answer = answer
                    break

            if not matched_answer:
                continue

            result_cell = sheet.cell(
                row=current_row,
                column=column,
                value=matched_answer.get("value", "") or ""
            )

            result_cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        current_row += 1
                    
    if checklist.get("score_enabled"):
        score_row = current_row

        sheet.cell(
            row=score_row,
            column=1,
            value="合計点"
        )

        sheet.cell(
            row=score_row,
            column=1
        ).font = Font(bold=True)

        for period_result in period_results:
            total_score = 0

            for answer in period_result.get("answers", []):
                try:
                    total_score += float(answer.get("value") or 0)
                except (TypeError, ValueError):
                    pass

            if excel_display_mode == "day":
                try:
                    column = int(period_result.get("day", 0)) + 1
                except (TypeError, ValueError):
                    continue

            elif excel_display_mode == "month":
                try:
                    column = int(period_result.get("month", 0)) + 1
                except (TypeError, ValueError):
                    continue

            else:
                try:
                    period_year = int(period_result.get("year", 0))
                except (TypeError, ValueError):
                    continue

                year_offset = period_year - base_year

                if year_offset < 0 or year_offset >= 5:
                    continue

                column = year_offset + 2

            sheet.cell(
                row=score_row,
                column=column,
                value=int(total_score)
            )

        current_row += 1
                                    
    # 点検実施者
    inspector_row = current_row

    sheet.cell(
        row=inspector_row,
        column=1,
        value="点検実施者"
    )

    for period_result in period_results:

        if excel_display_mode == "day":
            try:
                period_value = int(
                    period_result.get("day", 0)
                )
            except (TypeError, ValueError):
                continue

            if period_value < 1 or period_value > 31:
                continue

            column = period_value + 1

        elif excel_display_mode == "month":
            try:
                period_value = int(
                    period_result.get("month", 0)
                )
            except (TypeError, ValueError):
                continue

            if period_value < 1 or period_value > 12:
                continue

            column = period_value + 1

        else:
            try:
                period_year = int(
                    period_result.get("year", 0)
                )
            except (TypeError, ValueError):
                continue

            year_offset = period_year - base_year

            if year_offset < 0 or year_offset >= 5:
                continue

            column = year_offset + 2

        inspector_cell = sheet.cell(
            row=inspector_row,
            column=column,
            value=period_result.get("checked_by", "") or ""
        )

        inspector_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    current_row += 1     

    # 承認欄
    approval_items = [
        item
        for item in checklist.get("items", [])
        if item.get("item_type") == "approval"
    ]

    for approval_index, approval_item in enumerate(approval_items):
        approval_row = current_row
        
        sheet.row_dimensions[approval_row].height = 28

        sheet.cell(
            row=approval_row,
            column=1,
            value=approval_item.get("approval_label", "") or "承認"
        )

        for period_result in period_results:
            approvals = period_result.get("approvals", [])

            if approval_index >= len(approvals):
                continue

            approval = approvals[approval_index]

            if not approval.get("approved_by"):
                continue

            if excel_display_mode == "day":
                try:
                    column = int(period_result.get("day", 0)) + 1
                    
                except (TypeError, ValueError):
                    continue

            elif excel_display_mode == "month":
                try:
                    column = int(period_result.get("month", 0)) + 1
                except (TypeError, ValueError):
                    continue

            else:
                try:
                    period_year = int(period_result.get("year", 0))
                except (TypeError, ValueError):
                    continue

                year_offset = period_year - base_year

                if year_offset < 0 or year_offset >= 5:
                    continue

                column = year_offset + 2

            sheet.cell(
                row=approval_row,
                column=column,
                value=approval.get("approved_by", "")
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        current_row += 1

    # 表全体の罫線・配置
    thin = Side(
        style="thin",
        color="000000"
    )

    table_end_row = current_row - 1

    for row_number in range(5, table_end_row + 1):

        # カテゴリ行は横一列にして日付マスの縦線を付けない
        if row_number in category_rows:
            no_border = Side(style=None)

            for category_column in range(
                1,
                title_end_column + 1
            ):
                category_cell = sheet.cell(
                    row=row_number,
                    column=category_column
                )

                category_cell.border = Border(
                    left=(
                        thin
                        if category_column == 1
                        else no_border
                    ),
                    right=(
                        thin
                        if category_column == title_end_column
                        else no_border
                    ),
                    top=thin,
                    bottom=thin
                )

            continue

        for column_number in range(1, title_end_column + 1):
            
            cell = sheet.cell(
                row=row_number,
                column=column_number
            )

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            if column_number >= 2:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

    # ヘッダー装飾
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9E1F2"
    )

    for row_number in range(header_row, weekday_row + 1):
        for column_number in range(1, title_end_column + 1):
            header_cell = sheet.cell(
                row=row_number,
                column=column_number
            )

            header_cell.fill = header_fill
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # 日次表の土日列を薄く色付け
    if excel_display_mode == "day":
        saturday_fill = PatternFill(
            fill_type="solid",
            fgColor="EAF2F8"
        )

        sunday_fill = PatternFill(
            fill_type="solid",
            fgColor="FCE8E6"
        )

        for day in range(1, days_in_month + 1):
            if day > days_in_month:
                continue

            column = day + 1

            weekday_index = datetime(
                year,
                month,
                day
            ).weekday()

            if weekday_index == 5:
                fill = saturday_fill
            elif weekday_index == 6:
                fill = sunday_fill
            else:
                continue

            for row_number in range(
                header_row,
                table_end_row + 1
            ):
                sheet.cell(
                    row=row_number,
                    column=column
                ).fill = fill

                                
    # 日次表の土日文字色を再設定
    if excel_display_mode == "day":
        for column_number in range(2, title_end_column + 1):

            weekday_cell = sheet.cell(
                row=weekday_row,
                column=column_number
            )

            if weekday_cell.value == "土":
                weekday_cell.font = Font(
                    bold=True,
                    color="0000FF"
                )

            elif weekday_cell.value == "日":
                weekday_cell.font = Font(
                    bold=True,
                    color="FF0000"
                )
                
    # Excelファイルをメモリ上に保存
    output = BytesIO()
    sheet.print_area = (
        f"A1:{sheet.cell(
            row=table_end_row,
            column=title_end_column
        ).coordinate}"
    )
    
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2
        
    workbook.save(output)
    output.seek(0)

    filename = (
        f"{checklist.get('name', '車両点検表')}_"
        f"{result_record.year}"
    )

    if excel_display_mode == "day":
        filename += f"_{int(result_record.month):02d}"

    filename += ".xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
                                                
@app.route("/vehicle/checklists/<int:index>/save-one", methods=["POST"])
def save_vehicle_checklist_one(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/vehicle/checklists")

    if session.get("role") != "itc":
        if checklist_record.company_code != session.get("company_code"):
            return redirect("/vehicle/checklists")

    checklist = checklist_to_dict(checklist_record)

    vehicle_id = request.form.get("vehicle_id")
    year = request.form.get("year", "")
    month = request.form.get("month", "")
    day = request.form.get("day", "")

    if checklist.get("frequency_unit") == "year":
        month = "01"
        day = "01"
    elif checklist.get("display_type") == "month":
        month = str(month).zfill(2)
        day = str(day).zfill(2)
    else:
        month = str(month).zfill(2)
        day = "01"

    active_day = request.form.get("active_day")

    item_no = request.form.get("item_no")
    category = request.form.get("category")
    content = request.form.get("content")
    criteria = request.form.get("criteria")
    value = request.form.get("value")

    result_record = VehicleChecklistResult.query.filter_by(
        company_code=checklist_record.company_code,
        checklist_id=checklist_record.id,
        vehicle_id=vehicle_id,
        year=year,
        month=month,
        day=day
    ).first()
    
    if not result_record:
        result_record = VehicleChecklistResult(
            company_code=checklist_record.company_code,
            checklist_id=checklist_record.id,
            vehicle_id=vehicle_id,
            year=year,
            month=month,
            day=day,
            checked_by=session.get("name"),
            checked_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
            status="入力中",
            approved_by="",
            approved_date="",
            reject_reason="",
            answers_json="[]"
        )

        db.session.add(result_record)

    answers = json.loads(result_record.answers_json or "[]")

    answer = None

    for a in answers:
        if str(a.get("item_no", "")) == str(item_no):
            answer = a
            break

    if not answer:
        answer = {
            "item_no": item_no,
            "category": category,
            "content": content,
            "criteria": criteria,
            "value": "",
            "comment": "",
            "files": []
        }
        answers.append(answer)

    answer["value"] = value
    answer["category"] = category
    answer["content"] = content
    answer["criteria"] = criteria

    result_record.checked_by = session.get("name")
    result_record.checked_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    result_record.status = "入力中"
    result_record.approved_by = ""
    result_record.approved_date = ""
    result_record.answers_json = json.dumps(answers, ensure_ascii=False)

    db.session.commit()

    return redirect(
        f"/vehicle/checklists/{checklist_record.id}?vehicle_id={vehicle_id}&year={year}&month={month}&active_day={active_day}"
    )

@app.route("/vehicle/checklists/<int:index>/save-detail", methods=["POST"])
def save_vehicle_checklist_detail(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/vehicle/checklists")

    if session.get("role") != "itc":
        if checklist_record.company_code != session.get("company_code"):
            return redirect("/vehicle/checklists")

    checklist = checklist_to_dict(checklist_record)

    vehicle_id = request.form.get("vehicle_id")
    year = request.form.get("year", "")
    month = request.form.get("month", "")
    day = request.form.get("day", "")

    if checklist.get("frequency_unit") == "year":
        month = "01"
        day = "01"
    elif checklist.get("display_type") == "month":
        month = str(month).zfill(2)
        day = str(day).zfill(2)
    else:
        month = str(month).zfill(2)
        day = "01"

    content = request.form.get("content")
    item_no = request.form.get("item_no")
    active_day = request.form.get("active_day")
    category = request.form.get("category")
    criteria = request.form.get("criteria")
    comment = request.form.get("comment")

    result_record = VehicleChecklistResult.query.filter_by(
        company_code=checklist_record.company_code,
        checklist_id=checklist_record.id,
        vehicle_id=vehicle_id,
        year=year,
        month=month,
        day=day
    ).first()

    if not result_record:
        result_record = VehicleChecklistResult(
            company_code=session.get("company_code"),
            checklist_id=checklist_record.id,
            vehicle_id=vehicle_id,
            year=year,
            month=month,
            day=day,
            checked_by=session.get("name"),
            checked_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
            status="入力中",
            approved_by="",
            approved_date="",
            reject_reason="",
            answers_json="[]"
        )
        db.session.add(result_record)

    answers = json.loads(result_record.answers_json or "[]")

    answer = None

    for a in answers:
        if str(a.get("item_no", "")) == str(item_no):
            answer = a
            break

    if not answer:
        answer = {
            "category": category,
            "item_no": item_no,
            "content": content,
            "criteria": criteria,
            "value": "",
            "comment": "",
            "files": []
        }
        answers.append(answer)

    uploaded_files = request.files.getlist("files")
    answer.setdefault("files", [])

    for file in uploaded_files:
        filename = save_uploaded_file(file)

        if filename:
            answer["files"].append(filename)

    answer["comment"] = comment
    answer["category"] = category
    answer["content"] = content
    answer["criteria"] = criteria

    result_record.checked_by = session.get("name")
    result_record.checked_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    result_record.status = "入力中"
    result_record.approved_by = ""
    result_record.approved_date = ""
    result_record.answers_json = json.dumps(answers, ensure_ascii=False)

    if request.form.get("patrol_link") == "1":
        db.session.add(VehiclePatrol(
            company_code=checklist_record.company_code,
            vehicle_id=vehicle_id,
            occurred_date=f"{year}-{month}-{day}",
            category="点検指摘",
            priority="中",
            content=content,
            cause="",
            temporary_action=comment,
            repair_content="",
            status="未対応",
            repair_date="",
            repair_person="",
            repair_time="",
            parts="",
            cost=""
        ))

    notify_mentions(
        comment,
        f"/vehicle/checklists/{checklist_record.id}?vehicle_id={vehicle_id}&year={year}&month={month}&active_day={active_day}"
    )

    db.session.commit()

    return redirect(
        f"/vehicle/checklists/{checklist_record.id}?vehicle_id={vehicle_id}&year={year}&month={month}&active_day={active_day}"
    )

@app.route("/vehicle/checklists/<int:index>/complete", methods=["POST"])
def complete_vehicle_checklist(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/vehicle/checklists")

    if session.get("role") != "itc":
        if checklist_record.company_code != session.get("company_code"):
            return redirect("/vehicle/checklists")

    vehicle_id = request.form.get("vehicle_id")
    year = request.form.get("year", "")
    month = request.form.get("month", "")
    day = request.form.get("day", "")
    active_day = request.form.get("active_day", "")

    checklist = checklist_to_dict(checklist_record)

    if checklist.get("frequency_unit") == "year":
        month = "01"
        day = "01"
    elif checklist.get("display_type") == "month":
        month = str(month).zfill(2)
        day = str(day).zfill(2)
    else:
        month = str(month).zfill(2)
        day = "01"

    result_record = VehicleChecklistResult.query.filter_by(
        company_code=checklist_record.company_code,
        checklist_id=checklist_record.id,
        vehicle_id=vehicle_id,
        year=year,
        month=month,
        day=day
    ).first()

    if not result_record:
        return redirect(
            f"/vehicle/checklists/{checklist_record.id}"
            f"?vehicle_id={vehicle_id}"
            f"&year={year}"
            f"&month={month}"
            f"&active_day={active_day}"
        )

    result_record.status = "承認待ち"
    result_record.checked_by = session.get("name")
    result_record.checked_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    result_record.approved_by = ""
    result_record.approved_date = ""
    
    notify_users = list(dict.fromkeys(
        name.strip()
        for name in request.form.getlist("notify_users")
        if name.strip()
    ))

    result_record.notify_users_json = json.dumps(
        notify_users,
        ensure_ascii=False
    )

    db.session.commit()

    notification_link = (
        f"/vehicle/checklists/{checklist_record.id}"
        f"?vehicle_id={vehicle_id}"
        f"&year={year}"
        f"&month={month}"
        f"&active_day={active_day}"
    )

    for target_user in set(notify_users):
        add_notification(
            target_user,
            "車両点検完了のお知らせ",
            (
                f"{vehicle_id} の「{checklist_record.name}」が"
                f"点検完了しました。"
            ),
            notification_link
        )

    return redirect(
        f"/vehicle/checklists/{checklist_record.id}"
        f"?vehicle_id={vehicle_id}"
        f"&year={year}"
        f"&month={month}"
        f"&active_day={active_day}"
    )
    
@app.route("/vehicle/checklists/<int:index>/new", methods=["GET", "POST"])
def new_vehicle_checklist_result(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/vehicle/checklists")

    if session.get("role") != "itc":
        if checklist_record.company_code != session.get("company_code"):
            return redirect("/vehicle/checklists")

    checklist = checklist_to_dict(checklist_record)

    if checklist["target"] != "車両管理":
        return redirect("/vehicle/checklists")

    if request.method == "POST":
        vehicle_id = request.form.get("vehicle_id")
        year = request.form.get("year")
        month = request.form.get("month")
        day = request.form.get("day")

        answers = []
        answer_index = 0

        for item in checklist["items"]:
            if item.get("item_type") == "approval":
                continue

            file_names = []

            uploaded_files = request.files.getlist(f"files_{answer_index}")

            for file in uploaded_files:
                filename = save_uploaded_file(file)

                if filename:
                    file_names.append(filename)

            answers.append({
                "item_no": answer_index,
                "category": item.get("category", ""),
                "content": item.get("content", ""),
                "criteria": item.get("criteria", ""),
                "value": request.form.get(f"answer_{answer_index}"),
                "comment": request.form.get(f"comment_{answer_index}"),
                "files": file_names
            })

            answer_index += 1
        approvals = []

        for item in checklist.get("items", []):
            if item.get("item_type") != "approval":
                continue

            approvals.append({
                "label": item.get("approval_label", ""),
                "allow_general": item.get("approval_allow_general", False),
                "approved_by": "",
                "approved_date": ""
            })
            
        result = VehicleChecklistResult(
            company_code=checklist_record.company_code,
            checklist_id=checklist_record.id,
            vehicle_id=vehicle_id,
            year=year,
            month=str(month).zfill(2),
            day=str(day).zfill(2),
            checked_by=session.get("name"),
            checked_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
            status="承認待ち",
            approved_by="",
            approved_date="",
            reject_reason="",
            approvals_json=json.dumps(
                approvals,
                ensure_ascii=False
            ),
            answers_json=json.dumps(answers, ensure_ascii=False)
        )

        db.session.add(result)
        db.session.commit()

        return redirect(
            f"/vehicle/checklists/{checklist_record.id}?vehicle_id={vehicle_id}&year={year}&month={str(month).zfill(2)}"
        )

    return render_template(
        "vehicle_checklist_form.html",
        checklist=checklist,
        checklist_index=checklist_record.id,
        mode="new",
        now_year=datetime.now().year,
        now_month=datetime.now().month,
        now_day=datetime.now().day
    )

@app.route("/safety/checklists/<int:index>/new", methods=["GET", "POST"])
def new_safety_checklist_result(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/safety/checklists")

    if session.get("role") != "itc":
        if checklist_record.company_code != session.get("company_code"):
            return redirect("/safety/checklists")

    checklist = checklist_to_dict(checklist_record)

    if request.method == "POST":
        answers = []
        answer_index = 0

        target_user = request.form.get("target_user") or session.get("name")

        target_driver = Driver.query.filter_by(
            company_code=session.get("company_code"),
            name=target_user
        ).first()

        target_office = target_driver.office if target_driver else session.get("office")

        for item in checklist["items"]:
            if item.get("item_type") == "approval":
                continue

            value = request.form.get(f"answer_{answer_index}")
            comment = request.form.get(f"comment_{answer_index}")
            patrol_link = request.form.get(f"patrol_link_{answer_index}")

            file_names = []

            uploaded_files = request.files.getlist(f"files_{answer_index}")

            for file in uploaded_files:
                filename = save_uploaded_file(file)

                if filename:
                    file_names.append(filename)

            answers.append({
                "category": item.get("category", ""),
                "content": item.get("content", ""),
                "criteria": item.get("criteria", ""),
                "criteria_files": item.get("criteria_files", []),
                "value": value,
                "comment": comment,
                "files": file_names,
                "patrol_link": patrol_link == "1",
            })
            
            if patrol_link == "1":
                db.session.add(PatrolResult(
                    company_code=session.get("company_code"),
                    created_by_username=session.get("username"),
                    created_by_name=session.get("name"),
                    date=datetime.now().strftime("%Y-%m-%d"),
                    delivery_place="",
                    category="点検指摘",
                    content_type="安全",
                    target_type=request.form.get("target_type") or "user",
                    target_user=target_user,
                    office=target_office,
                    content=(
                        f"{item.get('category', '')}：{item.get('content', '')}"
                        f" / 評価：{value or '-'}"
                        f" / コメント：{comment or '-'}"
                    ),
                    files_json=json.dumps(file_names, ensure_ascii=False),
                    countermeasure="",
                    approval_status="未対応",
                    reject_reason=""
                ))

            answer_index += 1

        approvals = []

        for item in checklist.get("items", []):
            if item.get("item_type") != "approval":
                continue

            approvals.append({
                "label": item.get("approval_label", ""),
                "allow_general": item.get("approval_allow_general", False),
                "approved_by": "",
                "approved_date": "",
            })
        result = ChecklistResult(
            company_code=session.get("company_code"),
            checklist_id=checklist_record.id,
            target_type=request.form.get("target_type"),
            target_user=request.form.get("target_user"),
            target_vehicle=request.form.get("target_vehicle"),
            target_office=request.form.get("target_office"),
            checked_by=session.get("name"),
            checked_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
            approved_by="",
            approved_date="",
            reject_reason="",
            status="承認待ち",
            approvals_json=json.dumps(approvals, ensure_ascii=False),
            answers_json=json.dumps(answers, ensure_ascii=False)
        )

        db.session.add(result)
        db.session.commit()

        return redirect(f"/safety/checklists/{checklist_record.id}")

    criteria_list = []

    for item in checklist["items"]:
        if item.get("item_type") == "approval":
            continue

        criteria = item.get("criteria", "")
        if criteria and criteria not in criteria_list:
            criteria_list.append(criteria)

    return render_template(
        "checklist_result_form.html",
        checklist=checklist,
        index=checklist_record.id,
        criteria_list=criteria_list,
        drivers=drivers_for_current_company(),
        selected_vehicle=None,
        offices=offices_for_current_company()
    )

@app.route("/master/checklists/<int:index>/edit", methods=["GET", "POST"])
def edit_checklist(index):
    checklist_record = Checklist.query.get(index)

    if not checklist_record:
        return redirect("/master/checklists")

    if session.get("role") != "itc":
        if checklist_record.company_code != session.get("company_code"):
            return redirect("/master/checklists")

    checklist = checklist_to_dict(checklist_record)

    if request.method == "POST":
        item_categories = request.form.getlist("item_category")
        item_contents = request.form.getlist("item_content")
        input_types = request.form.getlist("input_type")
        item_types = request.form.getlist("item_type")
        approval_labels = request.form.getlist("approval_label")
        approval_allow_general_list = request.form.getlist("approval_allow_general")
        choices_list = request.form.getlist("choices")
        criteria_list = request.form.getlist("criteria")
        comment_required_list = request.form.getlist("comment_required")
        score_enabled = request.form.get("score_enabled") == "1"
        print_portrait = (
            request.form.get("target") == "車両管理"
            and request.form.get("print_portrait") == "1"
        )
        old_items = checklist.get("items", [])
        items = []

        for i in range(len(item_types)):
            item_type = item_types[i]
            
            if item_type == "inspector":
                items.append({
                    "item_type": "inspector",
                })
                continue

            if item_type == "approval":
                label = ""

                if i < len(approval_labels):
                    label = approval_labels[i]

                items.append({
                    "item_type": "approval",
                    "approval_label": label,
                    "approval_allow_general": str(i) in approval_allow_general_list,
                    "criteria_files": [],
                })

                continue

            if i >= len(item_contents):
                continue

            if not item_contents[i]:
                continue

            choices = []

            if input_types[i] == "select":
                choices = [
                    choice.strip()
                    for choice in choices_list[i].split(",")
                    if choice.strip()
                ]

            criteria_files = []

            if i < len(old_items):
                criteria_files = list(old_items[i].get("criteria_files", []))

            for file in request.files.getlist(f"criteria_files_{i}"):
                filename = save_uploaded_file(file)

                if filename:
                    criteria_files.append(filename)

            items.append({
                "item_type": "check",
                "category": item_categories[i],
                "content": item_contents[i],
                "input_type": input_types[i],
                "choices": choices,
                "criteria": criteria_list[i],
                "criteria_files": criteria_files,
                "comment_required": str(i) in comment_required_list,
                "score_enabled": score_enabled,
            })

        checklist_record.name = request.form.get("name")
        checklist_record.target = request.form.get("target")

        if request.form.get("target") == "車両管理":
            checklist_record.frequency_value = request.form.get("frequency_value")
            checklist_record.frequency_unit = request.form.get("frequency_unit")
            checklist_record.display_type = request.form.get("display_type")
            checklist_record.print_portrait = print_portrait
        else:
            checklist_record.frequency_value = ""
            checklist_record.frequency_unit = ""
            checklist_record.display_type = ""
            checklist_record.print_portrait = False

        checklist_record.items_json = json.dumps(items, ensure_ascii=False)

        db.session.commit()

        return redirect("/master/checklists")

    return render_template(
        "checklist_form.html",
        checklist=checklist,
        index=checklist_record.id,
        mode="edit"
    )


@app.route("/master/checklists/<int:index>/delete", methods=["POST"])
def delete_checklist(index):
    checklist = Checklist.query.get(index)

    if not checklist:
        return redirect("/master/checklists")

    if session.get("role") != "itc":
        if checklist.company_code != session.get("company_code"):
            return redirect("/master/checklists")

    db.session.delete(checklist)
    db.session.commit()

    return redirect("/master/checklists")


@app.route(
    "/safety/checklist-results/<int:result_index>/approve/<int:approval_index>",
    methods=["POST"]
)
def approve_checklist_result(result_index, approval_index):
    result_record = ChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/safety/checklists")

    result = checklist_result_to_dict(result_record)

    approvals = result.get("approvals", [])
    if not approvals:
        checklist_record = Checklist.query.get(result_record.checklist_id)

        if checklist_record:
            checklist = checklist_to_dict(checklist_record)

            for item in checklist.get("items", []):
                if item.get("item_type") != "approval":
                    continue

                approvals.append({
                    "label": item.get("approval_label", ""),
                    "allow_general": item.get("approval_allow_general", False),
                    "approved_by": "",
                    "approved_date": "",
                })

    if approval_index < 0 or approval_index >= len(approvals):
        return redirect(f"/safety/checklist-results/{result_index}")

    approval = approvals[approval_index]

    if not can_approve_checklist_result(result, approval):
        return redirect(f"/safety/checklist-results/{result_index}")

    approval["approved_by"] = session.get("name")
    approval["approved_date"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    result_record.approvals_json = json.dumps(
        approvals,
        ensure_ascii=False
    )

    result_record.reject_reason = ""

    all_approved = all(
        item.get("approved_by")
        for item in approvals
    )

    if approvals and all_approved:
        result_record.status = "承認済み"
        result_record.approved_by = session.get("name")
        result_record.approved_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    else:
        result_record.status = "承認待ち"

    db.session.commit()

    return redirect(f"/safety/checklist-results/{result_record.id}")

@app.route("/safety/checklist-results/<int:result_index>/reject", methods=["POST"])
def reject_checklist_result(result_index):
    result_record = ChecklistResult.query.get(result_index)

    if not result_record:
        return redirect("/safety/checklists")

    result = checklist_result_to_dict(result_record)

    if not can_reject_checklist_result(result):
        return redirect(f"/safety/checklist-results/{result_index}")

    reject_reason = request.form.get("reject_reason", "")

    result_record.status = "差し戻し"
    result_record.approved_by = session.get("name")
    result_record.approved_date = ""
    result_record.reject_reason = reject_reason

    notify_users = set()

    if result_record.checked_by:
        notify_users.add(result_record.checked_by)

    if result_record.target_user:
        notify_users.add(result_record.target_user)

    for target_user in notify_users:
        add_notification(
            target_user,
            "チェックリストが差し戻されました",
            reject_reason or "チェックリストが差し戻されました。",
            f"/safety/checklist-results/{result_record.id}"
        )

    db.session.commit()

    return redirect(f"/safety/checklist-results/{result_record.id}")

def init_db():
    with app.app_context():
        db.create_all()

        if not Company.query.filter_by(company_code="ITC").first():
            db.session.add(Company(
                company_code="ITC",
                company_name="ITC",
                vehicle_limit=9999,
                active=True
            ))
            db.session.commit()

        if not User.query.filter_by(company_code="ITC", username="itc").first():
            db.session.add(User(
                company_code="ITC",
                username="itc",
                password=generate_password_hash("itc123"),
                role="itc",
                name="ITC管理者",
                office="ITC",
                favorite_vehicles_json="[]"
            ))

        default_content_types = [
            "落下",
            "車両",
            "環境",
            "荷扱い",
            "ルール違反",
            "Good",
            "その他"
        ]

        for company in Company.query.all():
            for name in default_content_types:
                exists = PatrolContentType.query.filter_by(
                    company_code=company.company_code,
                    name=name
                ).first()

                if not exists:
                    db.session.add(PatrolContentType(
                        company_code=company.company_code,
                        name=name
                    ))

        db.session.commit()


init_db()


with app.app_context():

    columns = [
        ("chassis_number", "VARCHAR(100)"),
        ("model_code", "VARCHAR(100)"),
        ("first_registration_date", "VARCHAR(20)"),
        ("manufacturer", "VARCHAR(100)"),
        ("body_type", "VARCHAR(100)"),
        ("gross_vehicle_weight", "INTEGER"),
        ("max_payload", "INTEGER"),
    ]

    inspector = inspect(db.engine)

    existing_columns = [
        column["name"]
        for column in inspector.get_columns("vehicle")
    ]

    for column_name, column_type in columns:
        if column_name not in existing_columns:
            db.session.execute(
                db.text(
                    f"ALTER TABLE vehicle "
                    f"ADD COLUMN {column_name} {column_type}"
                )
            )

    db.session.commit()
    
    checklist_columns = [
        ("notify_users_json", "TEXT"),
        (
            "print_portrait",
            "BOOLEAN NOT NULL DEFAULT FALSE"
        ),
    ]

    existing_checklist_columns = [
        column["name"]
        for column in inspector.get_columns("checklist")
    ]

    for column_name, column_type in checklist_columns:
        if column_name not in existing_checklist_columns:
            db.session.execute(
                db.text(
                    f"ALTER TABLE checklist "
                    f"ADD COLUMN {column_name} {column_type}"
                )
            )

    checklist_result_columns = [
        ("approvals_json", "TEXT"),
    ]

    existing_checklist_result_columns = [
        column["name"]
        for column in inspector.get_columns(
            "checklist_result"
        )
    ]

    for column_name, column_type in checklist_result_columns:
        if column_name not in existing_checklist_result_columns:
            db.session.execute(
                db.text(
                    f"ALTER TABLE checklist_result "
                    f"ADD COLUMN {column_name} {column_type}"
                )
            )

    db.session.commit()

    vehicle_checklist_result_columns = [
        ("notify_users_json", "TEXT"),
        ("approvals_json", "TEXT"),
    ]

    existing_vehicle_checklist_result_columns = [
        column["name"]
        for column in inspector.get_columns(
            "vehicle_checklist_result"
        )
    ]

    for column_name, column_type in vehicle_checklist_result_columns:
        if column_name not in existing_vehicle_checklist_result_columns:
            db.session.execute(
                db.text(
                    f"ALTER TABLE vehicle_checklist_result "
                    f"ADD COLUMN {column_name} {column_type}"
                )
            )

    db.session.commit()
    
if __name__ == "__main__":
    app.run(debug=False)